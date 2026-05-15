import streamlit as st
import pandas as pd
import numpy as np
from supabase import create_client
from datetime import datetime
import io
import os
import base64
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===================== إعدادات الصفحة =====================
st.set_page_config(
    page_title="نظام كاريتاس",
    layout="wide",
    page_icon="📊",
    initial_sidebar_state="collapsed"
)

# ===================== CSS الكامل =====================
st.markdown("""
<style>
[data-testid="collapsedControl"]    { display: none !important; }
[data-testid="stSidebar"]           { display: none !important; }
section[data-testid="stSidebarNav"] { display: none !important; }
#MainMenu        { visibility: hidden; }
footer           { visibility: hidden; }
header           { visibility: hidden; }
.stDeployButton  { display: none; }

html, body, .main, .block-container { direction: rtl; }

/* ======= هيدر ======= */
.sys-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 60%, #2563eb 100%);
    padding: 14px 28px;
    border-radius: 18px;
    margin-bottom: 24px;
    box-shadow: 0 8px 32px rgba(30,58,138,0.25);
}
.sys-header-logo { display:flex; align-items:center; gap:14px; }
.sys-header-logo img { height:52px; border-radius:10px; }
.sys-header-title { color:#fff; font-size:22px; font-weight:800; letter-spacing:-0.5px; }
.sys-header-sub   { color:#93c5fd; font-size:12px; margin-top:2px; }
.sys-header-user  { text-align:left; color:#bfdbfe; font-size:13px; line-height:1.7; }

/* ======= كروت البرامج ======= */
.prog-cards-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 24px;
    margin-bottom: 32px;
}
@media (max-width: 900px) {
    .prog-cards-grid { grid-template-columns: 1fr; }
}
.prog-card {
    background:#fff;
    border-radius:20px;
    padding:30px 26px 20px;
    box-shadow:0 4px 24px rgba(30,58,138,0.10);
    border:2px solid #e0e7ff;
    position:relative;
    overflow:hidden;
    cursor:pointer;
    transition: box-shadow 0.2s, transform 0.15s;
}
.prog-card:hover {
    box-shadow: 0 8px 32px rgba(30,58,138,0.18);
    transform: translateY(-2px);
}
.prog-card::before {
    content:'';
    position:absolute;
    top:0; right:0;
    width:6px; height:100%;
    background:linear-gradient(180deg,#2563eb,#06b6d4);
    border-radius:0 18px 18px 0;
}
.prog-card-icon { font-size:44px; margin-bottom:14px; display:block; }
.prog-card-name { font-size:20px; font-weight:800; color:#1e3a8a; margin-bottom:8px; }
.prog-card-desc { font-size:13px; color:#64748b; line-height:1.6; margin-bottom:18px; }
.prog-card-badge {
    display:inline-block;
    background:linear-gradient(90deg,#2563eb,#06b6d4);
    color:#fff; font-size:11px; font-weight:700;
    padding:4px 14px; border-radius:20px;
}
.prog-card-notif-badge {
    position:absolute;
    top:12px; left:14px;
    background:#ef4444;
    color:#fff;
    font-size:11px;
    font-weight:700;
    padding:2px 10px;
    border-radius:20px;
}

/* ======= KPIs ======= */
.kpi-card {
    background:#fff;
    border-radius:14px;
    padding:18px 20px;
    box-shadow:0 2px 12px rgba(30,58,138,0.09);
    border-top:4px solid #2563eb;
    text-align:center;
    margin-bottom:16px;
}
.kpi-val { font-size:22px; font-weight:800; color:#1e3a8a; }
.kpi-lbl { font-size:12px; color:#64748b; margin-top:4px; }

/* ======= كروت الأكواد ======= */
.code-card {
    background:#fff;
    border-radius:14px;
    padding:16px 18px;
    box-shadow:0 2px 12px rgba(30,58,138,0.09);
    border-top:4px solid #2563eb;
    text-align:center;
    margin-bottom:14px;
}
.code-card-name  { font-size:13px; font-weight:700; color:#2563eb; margin-bottom:6px; }
.code-card-count { font-size:22px; font-weight:800; color:#1e3a8a; }
.code-card-unit  { font-size:11px; color:#64748b; margin-top:2px; }
.code-card-amt   { font-size:15px; font-weight:700; color:#059669; margin-top:8px; }
.code-card-amt-l { font-size:11px; color:#64748b; }

/* ======= فلاتر ======= */
.filter-bar {
    background:#f1f5f9;
    border-radius:14px;
    padding:18px 20px;
    margin-bottom:20px;
    border:1px solid #e2e8f0;
}
.filter-bar-title { font-size:14px; font-weight:700; color:#1e3a8a; margin-bottom:12px; }

/* ======= عنوان قسم ======= */
.sec-title {
    font-size:17px; font-weight:800; color:#1e3a8a;
    border-right:4px solid #2563eb;
    padding-right:12px;
    margin:22px 0 12px;
}

/* ======= أزرار ======= */
.stButton > button {
    background:linear-gradient(90deg,#1e3a8a,#2563eb) !important;
    color:white !important;
    border:none !important;
    border-radius:10px !important;
    font-weight:700 !important;
}
.stButton > button:hover { opacity:0.88 !important; }
[data-testid="stDownloadButton"] > button {
    background:linear-gradient(90deg,#059669,#10b981) !important;
    color:white !important;
    border:none !important;
    border-radius:10px !important;
    font-weight:700 !important;
}
.stDataFrame { border:1px solid #e2e8f0; border-radius:12px; }

/* ======= بطاقة إشعار ======= */
.notif-card {
    background:#fff;
    border-radius:12px;
    padding:14px 18px;
    margin-bottom:10px;
    box-shadow:0 2px 10px rgba(30,58,138,0.08);
    border-right:4px solid #2563eb;
    position:relative;
}
.notif-card.unread { border-right-color:#ef4444; background:#fff7f7; }
.notif-card.approved { border-right-color:#059669; background:#f0fdf4; }
.notif-card.rejected { border-right-color:#dc2626; background:#fef2f2; }
.notif-card.edit_request { border-right-color:#f59e0b; background:#fffbeb; }
.notif-title { font-size:14px; font-weight:800; color:#1e3a8a; margin-bottom:4px; }
.notif-msg   { font-size:12px; color:#475569; line-height:1.5; }
.notif-time  { font-size:11px; color:#94a3b8; margin-top:4px; }
.notif-badge {
    display:inline-block;
    background:#ef4444;
    color:#fff;
    font-size:10px;
    font-weight:700;
    padding:2px 8px;
    border-radius:20px;
    position:absolute;
    top:10px;
    left:14px;
}

/* ======= موبايل ======= */
@media (max-width:640px) {
    .sys-header { flex-direction:column; gap:10px; text-align:center; padding:14px 16px; }
    .sys-header-user { text-align:center; }
    .prog-cards-grid { grid-template-columns: 1fr; }
}
</style>
""", unsafe_allow_html=True)

# ===================== Supabase =====================
URL = st.secrets["SUPABASE_URL"]
KEY = st.secrets["SUPABASE_KEY"]
supabase = create_client(URL, KEY)

PROGRAMS = {
    "reports": {
        "name": "سداد فوري & Opay",
        "icon": "💳",
        "desc": "عرض وتحليل بيانات السدادات — تقارير دقيقة ومتنوعة",
    },
    "installments": {
        "name": "الأقساط المستحقة",
        "icon": "📋",
        "desc": "متابعة الأقساط المستحقة مع بيانات الدفع من فوري و Opay",
    },
    "complaints": {
        "name": "شكاوى العملاء",
        "icon": "📝",
        "desc": "تسجيل ومتابعة شكاوى العملاء — إدخال وتعديل وموافقة إدارية",
    },
}

# ===================== دوال مساعدة عامة =====================

def get_logo_b64():
    for p in ["logo.png","images/logo.png","static/logo.png","assets/logo.png"]:
        if os.path.exists(p):
            with open(p,"rb") as f:
                return base64.b64encode(f.read()).decode()
    return None

def show_header(user=None):
    logo_b64  = get_logo_b64()
    full_name = user.get("full_name", "") if user else ""
    role      = user.get("role", "")      if user else ""

    if logo_b64:
        logo_part = (
            "<img src='data:image/png;base64," + logo_b64 +
            "' alt='logo' style='height:52px;border-radius:10px;'>"
        )
    else:
        logo_part = "<span style='font-size:32px'>📊</span>"

    if user:
        right_part = (
            "<div style='text-align:left;color:#bfdbfe;font-size:13px;line-height:1.8'>"
            "<div>👤 <strong style='color:#fff'>" + full_name + "</strong></div>"
            "<div style='font-size:11px;color:#7dd3fc'>" + role + "</div>"
            "</div>"
        )
    else:
        right_part = ""

    header_html = (
        "<div style='display:flex;align-items:center;justify-content:space-between;"
        "background:linear-gradient(135deg,#0f172a 0%,#1e3a8a 60%,#2563eb 100%);"
        "padding:14px 28px;border-radius:18px;margin-bottom:24px;"
        "box-shadow:0 8px 32px rgba(30,58,138,0.25);'>"
        "<div style='display:flex;align-items:center;gap:14px;'>"
        + logo_part +
        "<div>"
        "<div style='color:#fff;font-size:22px;font-weight:800;'>نظام كاريتاس</div>"
        "<div style='color:#93c5fd;font-size:12px;margin-top:2px;'>لوحة التقارير والمتابعة</div>"
        "</div>"
        "</div>"
        + right_part +
        "</div>"
    )
    st.markdown(header_html, unsafe_allow_html=True)

def check_login(u, p):
    res = supabase.table("app_users").select("*").eq("id",u).eq("password_hash",p).execute()
    return res.data[0] if res.data else None

def find_col(df, candidates):
    for c in candidates:
        if c in df.columns: return c
    return None

# ===================== جلب بيانات التقارير =====================

def fetch_reports_data():
    all_data, limit, offset = [], 1000, 0
    while True:
        res = supabase.table("all_payments_report").select("*").range(offset, offset+limit-1).execute()
        if not res.data: break
        all_data.extend(res.data)
        if len(res.data) < limit: break
        offset += limit
    df = pd.DataFrame(all_data)
    if not df.empty:
        df['تاريخ الدفع'] = pd.to_datetime(df['تاريخ الدفع'], dayfirst=True, errors='coerce')
    return df

def fetch_outstanding_data():
    all_data, limit, offset = [], 1000, 0
    while True:
        res = supabase.table("outstanding_with_payments").select("*").range(offset, offset+limit-1).execute()
        if not res.data: break
        all_data.extend(res.data)
        if len(res.data) < limit: break
        offset += limit
    return pd.DataFrame(all_data)

# ===================== Excel =====================

def thin_border():
    s = Side(border_style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def write_total_row(ws, total_row, cols, last_data_row):
    for ci in range(1, len(cols)+1):
        c = ws.cell(row=total_row, column=ci)
        c.fill=PatternFill("solid",fgColor="BFDBFE"); c.border=thin_border()
        c.alignment=Alignment(horizontal='center',vertical='center')
        c.font=Font(bold=True,color="1E3A8A",name="Arial",size=11)
    ws.cell(row=total_row, column=1).value = "✦ الإجمالي"
    if 'المبلغ' in cols:
        ci=cols.index('المبلغ')+1; cl=get_column_letter(ci)
        ws.cell(row=total_row,column=ci).value=f"=SUM({cl}3:{cl}{last_data_row})"
        ws.cell(row=total_row,column=ci).number_format='#,##0.00'
    ws.row_dimensions[total_row].height=26

def generate_reports_excel_single(df_display, sheet_title="التقرير", report_title="تقرير السدادات"):
    wb=Workbook(); ws=wb.active; ws.title=sheet_title[:31]
    ws.sheet_view.rightToLeft=True
    cols=list(df_display.columns); last_col=get_column_letter(len(cols))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].value=report_title; ws['A1'].font=Font(bold=True,size=14,color="1E3A8A",name="Arial")
    ws['A1'].fill=PatternFill("solid",fgColor="EFF6FF")
    ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=32
    for ci,h in enumerate(cols,1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
        c.fill=PatternFill("solid",fgColor="1E3A8A")
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=thin_border()
    ws.row_dimensions[2].height=24
    for ri,row in enumerate(df_display.itertuples(index=False),3):
        bg="F0F4FF" if ri%2==0 else "FFFFFF"
        for ci,val in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.fill=PatternFill("solid",fgColor=bg)
            c.alignment=Alignment(horizontal='center',vertical='center')
            c.font=Font(name="Arial",size=10); c.border=thin_border()
    ldr=2+len(df_display); write_total_row(ws,ldr+1,cols,ldr)
    cw={"اسم العميل":28,"الفرع":20,"مكان الدفع":16,"كود التحويلة":20,"رقم المرجع":18}
    for ci,col in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(ci)].width=cw.get(col,18)
    ws.freeze_panes="A3"
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

def generate_reports_excel_daily(df_display, original_df):
    wb=Workbook(); wb.remove(wb.active)
    DARK="1E3A8A"; LIGHT="EFF6FF"; WHT="FFFFFF"; TBG="BFDBFE"
    CODE_PALETTE = [
        ("DBEAFE","1E40AF"),("DCFCE7","166534"),("FEF9C3","854D0E"),("FCE7F3","9D174D"),
        ("EDE9FE","6B21A8"),("FFEDD5","9A3412"),("CFFAFE","0E7490"),("FEE2E2","991B1B"),
    ]
    all_codes = sorted(original_df['كود الخدمة'].dropna().unique().tolist())
    code_color_map = {c: CODE_PALETTE[i % len(CODE_PALETTE)] for i, c in enumerate(all_codes)}

    ws_sum = wb.create_sheet("ملخص يومي")
    ws_sum.sheet_view.rightToLeft = True
    sc = ["التاريخ", "كود الخدمة", "عدد الحركات", "إجمالي المبلغ (ج.م)"]
    nsc = len(sc)
    ws_sum.merge_cells(f'A1:{get_column_letter(nsc)}1')
    ws_sum['A1'].value = "ملخص يومي مفصّل — تقرير السدادات"
    ws_sum['A1'].font  = Font(bold=True, size=14, color=DARK, name="Arial")
    ws_sum['A1'].fill  = PatternFill("solid", fgColor=LIGHT)
    ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 32
    for ci, h in enumerate(sc, 1):
        c = ws_sum.cell(row=2, column=ci, value=h)
        c.font  = Font(bold=True, color=WHT, name="Arial", size=11)
        c.fill  = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
    ws_sum.row_dimensions[2].height = 24

    temp = original_df.copy()
    temp['_date'] = temp['تاريخ الدفع'].dt.date
    dates = sorted(temp['_date'].dropna().unique())
    cur_row = 3

    for d in dates:
        df_day = temp[temp['_date'] == d]
        day_start = cur_row
        for code in all_codes:
            df_code = df_day[df_day['كود الخدمة'] == code]
            if df_code.empty: continue
            bg_hex, fg_hex = code_color_map[code]
            cnt = len(df_code); amt = float(df_code['المبلغ'].sum())
            for ci in range(1, nsc + 1):
                c = ws_sum.cell(row=cur_row, column=ci)
                c.fill=PatternFill("solid",fgColor=bg_hex); c.font=Font(bold=True,color=fg_hex,name="Arial",size=10)
                c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
            ws_sum.cell(row=cur_row, column=1).value = str(d)
            ws_sum.cell(row=cur_row, column=2).value = code
            ws_sum.cell(row=cur_row, column=3).value = cnt
            ws_sum.cell(row=cur_row, column=4).value = amt
            ws_sum.cell(row=cur_row, column=4).number_format = '#,##0.00'
            cur_row += 1
        day_end = cur_row - 1
        for ci in range(1, nsc + 1):
            c = ws_sum.cell(row=cur_row, column=ci)
            c.fill=PatternFill("solid",fgColor="0F172A"); c.font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
            c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
        ws_sum.cell(row=cur_row, column=1).value = f"✦ إجمالي {d}"
        ws_sum.cell(row=cur_row, column=2).value = "الكل"
        ws_sum.cell(row=cur_row, column=3).value = f"=SUM(C{day_start}:C{day_end})"
        ws_sum.cell(row=cur_row, column=4).value = f"=SUM(D{day_start}:D{day_end})"
        ws_sum.cell(row=cur_row, column=4).number_format = '#,##0.00'
        ws_sum.cell(row=cur_row, column=3).font = Font(bold=True, color="BFDBFE", name="Arial", size=11)
        ws_sum.row_dimensions[cur_row].height = 24
        cur_row += 1

    data_end_sum = cur_row - 1
    for code in all_codes:
        bg_hex, fg_hex = code_color_map[code]
        for ci in range(1, nsc + 1):
            c = ws_sum.cell(row=cur_row, column=ci)
            c.fill=PatternFill("solid",fgColor=bg_hex); c.font=Font(bold=True,color=fg_hex,name="Arial",size=11)
            c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
        ws_sum.cell(row=cur_row, column=1).value = "✦ إجمالي كود"
        ws_sum.cell(row=cur_row, column=2).value = code
        ws_sum.cell(row=cur_row, column=3).value = f'=SUMIF(B3:B{data_end_sum},"{code}",C3:C{data_end_sum})'
        ws_sum.cell(row=cur_row, column=4).value = f'=SUMIF(B3:B{data_end_sum},"{code}",D3:D{data_end_sum})'
        ws_sum.cell(row=cur_row, column=4).number_format = '#,##0.00'
        ws_sum.row_dimensions[cur_row].height = 24
        cur_row += 1

    code_rows_start = cur_row - len(all_codes); code_rows_end = cur_row - 1
    for ci in range(1, nsc + 1):
        c = ws_sum.cell(row=cur_row, column=ci)
        c.fill=PatternFill("solid",fgColor="1E3A8A"); c.font=Font(bold=True,color="FFFFFF",name="Arial",size=12)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
    ws_sum.cell(row=cur_row, column=1).value = "✦ الإجمالي الكلي"
    ws_sum.cell(row=cur_row, column=2).value = "جميع الأكواد"
    ws_sum.cell(row=cur_row, column=3).value = f"=SUM(C{code_rows_start}:C{code_rows_end})"
    ws_sum.cell(row=cur_row, column=4).value = f"=SUM(D{code_rows_start}:D{code_rows_end})"
    ws_sum.cell(row=cur_row, column=4).number_format = '#,##0.00'
    ws_sum.row_dimensions[cur_row].height = 30
    for ci, w in enumerate([18, 18, 16, 22], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.freeze_panes = "A3"

    ws_br = wb.create_sheet("ملخص الفروع")
    ws_br.sheet_view.rightToLeft = True
    br_cols = ["الفرع", "التاريخ", "كود الخدمة", "عدد الحركات", "إجمالي المبلغ (ج.م)"]
    nbr = len(br_cols)
    ws_br.merge_cells(f'A1:{get_column_letter(nbr)}1')
    ws_br['A1'].value = "ملخص الفروع اليومي المفصّل"
    ws_br['A1'].font  = Font(bold=True, size=14, color=DARK, name="Arial")
    ws_br['A1'].fill  = PatternFill("solid", fgColor=LIGHT)
    ws_br['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_br.row_dimensions[1].height = 32
    for ci, h in enumerate(br_cols, 1):
        c = ws_br.cell(row=2, column=ci, value=h)
        c.font=Font(bold=True,color=WHT,name="Arial",size=11)
        c.fill=PatternFill("solid",fgColor=DARK)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
    ws_br.row_dimensions[2].height = 24

    all_branches = sorted(temp['branch_name'].dropna().unique().tolist()) if 'branch_name' in temp.columns else []
    br_row = 3
    for br in all_branches:
        df_br = temp[temp['branch_name'] == br]
        if df_br.empty: continue
        br_section_start = br_row
        for d in dates:
            df_day_br = df_br[df_br['_date'] == d]
            if df_day_br.empty: continue
            day_section_start = br_row
            for code in all_codes:
                df_dc = df_day_br[df_day_br['كود الخدمة'] == code]
                if df_dc.empty: continue
                bg_hex, fg_hex = code_color_map[code]
                cnt = len(df_dc); amt = float(df_dc['المبلغ'].sum())
                for ci in range(1, nbr + 1):
                    c = ws_br.cell(row=br_row, column=ci)
                    c.fill=PatternFill("solid",fgColor=bg_hex); c.font=Font(bold=True,color=fg_hex,name="Arial",size=10)
                    c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
                ws_br.cell(row=br_row,column=1).value=br; ws_br.cell(row=br_row,column=2).value=str(d)
                ws_br.cell(row=br_row,column=3).value=code; ws_br.cell(row=br_row,column=4).value=cnt
                ws_br.cell(row=br_row,column=5).value=amt; ws_br.cell(row=br_row,column=5).number_format='#,##0.00'
                br_row += 1
            day_end = br_row - 1
            for ci in range(1, nbr + 1):
                c = ws_br.cell(row=br_row, column=ci)
                c.fill=PatternFill("solid",fgColor="0F172A"); c.font=Font(bold=True,color="FFFFFF",name="Arial",size=10)
                c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
            ws_br.cell(row=br_row,column=1).value=br; ws_br.cell(row=br_row,column=2).value=f"✦ إجمالي {d}"
            ws_br.cell(row=br_row,column=3).value="الكل"
            ws_br.cell(row=br_row,column=4).value=f"=SUM(D{day_section_start}:D{day_end})"
            ws_br.cell(row=br_row,column=5).value=f"=SUM(E{day_section_start}:E{day_end})"
            ws_br.cell(row=br_row,column=5).number_format='#,##0.00'
            ws_br.row_dimensions[br_row].height=22; br_row += 1
        br_section_end = br_row - 1
        for ci in range(1, nbr + 1):
            c = ws_br.cell(row=br_row, column=ci)
            c.fill=PatternFill("solid",fgColor=TBG); c.font=Font(bold=True,color=DARK,name="Arial",size=11)
            c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
        ws_br.cell(row=br_row,column=1).value=f"✦ إجمالي {br}"; ws_br.cell(row=br_row,column=2).value="—"
        ws_br.cell(row=br_row,column=3).value="الكل"
        ws_br.cell(row=br_row,column=4).value=f'=SUMIF(C{br_section_start}:C{br_section_end},"الكل",D{br_section_start}:D{br_section_end})'
        ws_br.cell(row=br_row,column=5).value=f'=SUMIF(C{br_section_start}:C{br_section_end},"الكل",E{br_section_start}:E{br_section_end})'
        ws_br.cell(row=br_row,column=5).number_format='#,##0.00'; ws_br.row_dimensions[br_row].height=26; br_row+=1

    data_end_br = br_row - 1
    for code in all_codes:
        bg_hex, fg_hex = code_color_map[code]
        for ci in range(1, nbr + 1):
            c = ws_br.cell(row=br_row, column=ci)
            c.fill=PatternFill("solid",fgColor=bg_hex); c.font=Font(bold=True,color=fg_hex,name="Arial",size=11)
            c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
        ws_br.cell(row=br_row,column=1).value="✦ إجمالي كود"; ws_br.cell(row=br_row,column=2).value="—"
        ws_br.cell(row=br_row,column=3).value=code
        ws_br.cell(row=br_row,column=4).value=f'=SUMIF(C3:C{data_end_br},"{code}",D3:D{data_end_br})'
        ws_br.cell(row=br_row,column=5).value=f'=SUMIF(C3:C{data_end_br},"{code}",E3:E{data_end_br})'
        ws_br.cell(row=br_row,column=5).number_format='#,##0.00'; ws_br.row_dimensions[br_row].height=24; br_row+=1

    code_rows_start_br = br_row - len(all_codes); code_rows_end_br = br_row - 1
    for ci in range(1, nbr + 1):
        c = ws_br.cell(row=br_row, column=ci)
        c.fill=PatternFill("solid",fgColor="1E3A8A"); c.font=Font(bold=True,color="FFFFFF",name="Arial",size=12)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
    ws_br.cell(row=br_row,column=1).value="✦ الإجمالي الكلي"; ws_br.cell(row=br_row,column=2).value="—"
    ws_br.cell(row=br_row,column=3).value="جميع الأكواد"
    ws_br.cell(row=br_row,column=4).value=f"=SUM(D{code_rows_start_br}:D{code_rows_end_br})"
    ws_br.cell(row=br_row,column=5).value=f"=SUM(E{code_rows_start_br}:E{code_rows_end_br})"
    ws_br.cell(row=br_row,column=5).number_format='#,##0.00'; ws_br.row_dimensions[br_row].height=30
    for ci, w in enumerate([22, 16, 16, 16, 22], 1):
        ws_br.column_dimensions[get_column_letter(ci)].width = w
    ws_br.freeze_panes = "A3"

    def style_sheet_colored(ws, df_part, title_text):
        ws.sheet_view.rightToLeft = True
        cols = list(df_part.columns)
        ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
        ws['A1'].value=title_text; ws['A1'].font=Font(bold=True,size=13,color=DARK,name="Arial")
        ws['A1'].fill=PatternFill("solid",fgColor=LIGHT)
        ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[1].height=30
        for ci, h in enumerate(cols, 1):
            c=ws.cell(row=2,column=ci,value=h); c.font=Font(bold=True,color=WHT,name="Arial",size=10)
            c.fill=PatternFill("solid",fgColor=DARK); c.alignment=Alignment(horizontal='center',vertical='center')
            c.border=thin_border()
        ws.row_dimensions[2].height=22
        code_ci = None
        for ci, col in enumerate(cols, 1):
            if 'كود' in str(col) and 'خدمة' in str(col): code_ci=ci; break
            if str(col)=='كود الخدمة': code_ci=ci; break
        for ri, row in enumerate(df_part.itertuples(index=False), 3):
            row_code = None
            if code_ci: row_code = str(row[code_ci-1]) if len(row)>=code_ci else None
            if row_code and row_code in code_color_map:
                bg_hex, fg_hex = code_color_map[row_code]; use_alt=False
            else:
                bg_hex="F0F4FF" if ri%2==0 else "FFFFFF"; fg_hex="1E293B"; use_alt=True
            for ci, val in enumerate(row, 1):
                c=ws.cell(row=ri,column=ci,value=val)
                c.fill=PatternFill("solid",fgColor=bg_hex)
                c.font=Font(name="Arial",size=10,color=fg_hex,bold=(not use_alt))
                c.alignment=Alignment(horizontal='center',vertical='center'); c.border=thin_border()
        ldr=2+len(df_part); write_total_row(ws,ldr+1,cols,ldr); ws.freeze_panes="A3"
        cw={"اسم العميل":28,"الفرع":20,"كود الخدمة":16,"مكان الدفع":16,"كود التحويلة":20,"رقم المرجع":18,"وقت الدفع":12}
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width=cw.get(col,17)

    for d in dates:
        ddf=temp[temp['_date']==d].copy()
        dd=ddf.rename(columns={'client_code':'كود العميل','client_name':'اسم العميل','branch_name':'الفرع','CODE':'مكان الدفع','transaction_code':'كود التحويلة'})
        if 'تاريخ الدفع' in dd.columns: dd['تاريخ الدفع']=dd['تاريخ الدفع'].dt.strftime('%Y-%m-%d')
        drop=[c for c in dd.columns if c.startswith('_') or c=='id']
        dd=dd.drop(columns=drop,errors='ignore')
        if 'كود الخدمة' in dd.columns: dd=dd.sort_values('كود الخدمة').reset_index(drop=True)
        ws_day=wb.create_sheet(str(d)[:31]); style_sheet_colored(ws_day,dd,f"تقرير سدادات يوم {d}")

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

def generate_outstanding_excel(df, title="تقرير الأقساط المستحقة"):
    wb=Workbook(); ws=wb.active; ws.title="الأقساط المستحقة"
    ws.sheet_view.rightToLeft=True
    cols=list(df.columns); last_col=get_column_letter(len(cols))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].value=title; ws['A1'].font=Font(bold=True,size=16,color="1E3A8A",name="Arial")
    ws['A1'].fill=PatternFill("solid",fgColor="EFF6FF")
    ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=35
    headers=['اسم المسؤول','كود الفرع','اسم الفرع','تاريخ استحقاق القسط','تاريخ حالة القسط',
             'كود العميل','اسم العميل','الرقم القومي','رقم القرض','حالة القسط','قيمة القسط',
             'نوع الفاتورة','تاريخ التحويل','وقت التحويل','مبلغ فوري','رقم حساب الفوترة',
             'رقم تحويل فوري','الرقم المرجعي','مبلغ Opay','تاريخ الدفع Opay','وقت الدفع Opay']
    for ci,h in enumerate(headers[:len(cols)],1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
        c.fill=PatternFill("solid",fgColor="1E3A8A")
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=thin_border()
    ws.row_dimensions[2].height=30
    for ri,row in enumerate(df.itertuples(index=False),3):
        for ci,val in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=val)
            c.alignment=Alignment(horizontal='center',vertical='center')
            c.font=Font(name="Arial",size=10); c.border=thin_border()
            st_=str(row[9]) if len(row)>9 else ""
            if "مسدد جزئي" in st_:
                c.fill=PatternFill("solid",fgColor="FFCCCC"); c.font=Font(color="9C0006",bold=True)
            elif "غير مدفوع" in st_ or st_=="":
                c.fill=PatternFill("solid",fgColor="FFE699"); c.font=Font(color="7F4A00")
            else:
                c.fill=PatternFill("solid",fgColor="E2EFDA"); c.font=Font(color="375623")
    cw={'اسم العميل':25,'اسم الفرع':20,'اسم المسؤول':18,'الرقم القومي':15}
    for ci,col in enumerate(headers[:len(cols)],1):
        ws.column_dimensions[get_column_letter(ci)].width=cw.get(col,15)
    ws.freeze_panes="A3"
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ===================================================================
# ===================== دوال شكاوى العملاء ========================
# ===================================================================

def clean_json_value(value):
    """تحويل NaN و Infinity إلى None أو قيم صالحة لـ JSON"""
    if value is None:
        return None
    if isinstance(value, float):
        if np.isnan(value) or np.isinf(value):
            return None
    if isinstance(value, dict):
        return {k: clean_json_value(v) for k, v in value.items()}
    if isinstance(value, list):
        return [clean_json_value(v) for v in value]
    return value

def generate_complaint_number():
    try:
        year = datetime.now().year
        res = supabase.table("customer_complaints")\
            .select("complaint_number")\
            .like("complaint_number", f"شكوى-{year}-%")\
            .execute()
        existing = res.data or []
        nums = []
        for r in existing:
            try:
                n = int(r['complaint_number'].split('-')[-1])
                nums.append(n)
            except:
                pass
        seq = max(nums) + 1 if nums else 1
        return f"شكوى-{year}-{str(seq).zfill(5)}"
    except:
        import time
        return f"شكوى-{int(time.time())}"

def fetch_all_complaints(is_admin, user_branches):
    try:
        res = supabase.table("customer_complaints")\
            .select("*").order("submission_date", desc=True).execute()
        df = pd.DataFrame(res.data or [])
        if df.empty:
            return df
        if not is_admin and user_branches:
            df = df[df['branch_name'].isin(user_branches)]
        return df
    except Exception as e:
        st.error(f"خطأ في جلب الشكاوى: {e}")
        return pd.DataFrame()

def insert_new_complaint(data, user):
    try:
        res = supabase.table("customer_complaints").insert(data).execute()
        if res.data:
            comp = res.data[0]
            for admin_id in get_admin_user_ids():
                create_complaint_notification(
                    admin_id, "📝 شكوى جديدة",
                    f"تم تسجيل شكوى {comp['complaint_number']} من فرع {comp['branch_name']} — الشاكي: {comp['complainant_name']}",
                    "new_complaint", complaint_id=comp['id']
                )
            return True, comp
        return False, "لم يتم الحفظ"
    except Exception as e:
        return False, str(e)

def get_admin_user_ids():
    try:
        res = supabase.table("app_users").select("id").eq("role", "admin").execute()
        return [r['id'] for r in (res.data or [])]
    except:
        return []

def create_complaint_notification(user_id, title, message, notif_type,
                                   complaint_id=None, edit_req_id=None):
    try:
        data = {"user_id": user_id, "title": title, "message": message,
                "notification_type": notif_type, "is_read": False}
        if complaint_id:
            data["complaint_id"] = str(complaint_id)
        if edit_req_id:
            data["edit_request_id"] = str(edit_req_id)
        supabase.table("complaint_notifications").insert(data).execute()
    except:
        pass

def submit_edit_request_db(complaint_id, complaint_number, branch_name, changes, user):
    try:
        # تنظيف التغييرات من أي قيم NaN
        cleaned_changes = clean_json_value(changes)
        
        data = {
            "complaint_id": str(complaint_id),
            "complaint_number": complaint_number,
            "branch_name": branch_name,
            "requested_by": user.get('id'),
            "requester_name": user.get('full_name'),
            "changes": cleaned_changes,
            "status": "pending"
        }
        res = supabase.table("complaint_edit_requests").insert(data).execute()
        if res.data:
            req_id = res.data[0]['id']
            for admin_id in get_admin_user_ids():
                create_complaint_notification(
                    admin_id, "✏️ طلب تعديل شكوى",
                    f"طلب {user.get('full_name')} تعديل الشكوى رقم {complaint_number} — يرجى المراجعة",
                    "edit_request", complaint_id=complaint_id, edit_req_id=req_id
                )
            return True, req_id
        return False, "لم يتم الحفظ"
    except Exception as e:
        return False, str(e)

def fetch_edit_requests_db(is_admin=False, user_id=None, status_filter=None):
    try:
        q = supabase.table("complaint_edit_requests")\
            .select("*").order("requested_at", desc=True)
        if not is_admin and user_id:
            q = q.eq("requested_by", user_id)
        if status_filter and status_filter != "الكل":
            q = q.eq("status", status_filter)
        res = q.execute()
        return pd.DataFrame(res.data or [])
    except Exception as e:
        return pd.DataFrame()

def process_edit_request_db(request_id, complaint_id, changes, action,
                              admin_user, admin_note, requester_id):
    try:
        # تنظيف admin_note من NaN
        if admin_note and pd.isna(admin_note):
            admin_note = ""
        
        supabase.table("complaint_edit_requests").update({
            "status": action,
            "reviewed_by": admin_user.get('id'),
            "reviewed_by_name": admin_user.get('full_name'),
            "reviewed_at": datetime.now().isoformat(),
            "admin_note": admin_note or ""
        }).eq("id", request_id).execute()

        if action == "approved":
            update_data = {"updated_at": datetime.now().isoformat()}
            for field, vals in changes.items():
                if field != "_note" and isinstance(vals, dict):
                    new_val = vals.get('new')
                    # تنظيف القيمة الجديدة من NaN
                    if pd.isna(new_val):
                        new_val = None
                    update_data[field] = new_val
            supabase.table("customer_complaints").update(update_data)\
                .eq("id", str(complaint_id)).execute()
            create_complaint_notification(
                requester_id, "✅ تمت الموافقة على طلب التعديل",
                f"وافق {admin_user.get('full_name')} على تعديل الشكوى. تم تحديث البيانات.",
                "approved", complaint_id=complaint_id, edit_req_id=request_id
            )
        else:
            create_complaint_notification(
                requester_id, "❌ تم رفض طلب التعديل",
                f"رفض {admin_user.get('full_name')} طلب التعديل. السبب: {admin_note or 'لم يُذكر'}",
                "rejected", complaint_id=complaint_id, edit_req_id=request_id
            )
        return True
    except Exception as e:
        return False

def fetch_notifications_db(user_id):
    try:
        res = supabase.table("complaint_notifications")\
            .select("*").eq("user_id", user_id)\
            .order("created_at", desc=True).limit(50).execute()
        return pd.DataFrame(res.data or [])
    except:
        return pd.DataFrame()

def count_unread_notifications(user_id):
    try:
        res = supabase.table("complaint_notifications")\
            .select("id").eq("user_id", user_id).eq("is_read", False).execute()
        return len(res.data) if res.data else 0
    except:
        return 0

def mark_all_notifications_read(user_id):
    try:
        supabase.table("complaint_notifications")\
            .update({"is_read": True})\
            .eq("user_id", user_id).eq("is_read", False).execute()
    except:
        pass

def generate_complaints_excel(df):
    wb = Workbook(); ws = wb.active; ws.title = "سجل الشكاوى"
    ws.sheet_view.rightToLeft = True

    headers = [
        "رقم قيد الشكوى", "تاريخ التقديم", "الفرع",
        "اسم مقدم الشكوى", "صفة مقدم الشكوى", "رقم بطاقة العميل",
        "رقم الهاتف", "طريقة الاستقبال", "موجز الشكوى",
        "المستندات المرفقة", "تاريخ إبلاغ العميل",
        "طريقة الرد", "موجز نتيجة الفحص",
        "موقف الشكوى النهائي", "مبررات الرفض", "المدخل"
    ]
    db_cols = [
        "complaint_number", "submission_date", "branch_name",
        "complainant_name", "complainant_role", "client_card_number",
        "client_phone", "reception_method", "complaint_summary",
        "attached_documents", "notification_date",
        "response_method", "investigation_summary",
        "final_status", "rejection_justification", "created_by_name"
    ]

    last_col = get_column_letter(len(headers))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].value = f"سجل شكاوى العملاء — تقرير (ج.م.ص./10)"
    ws['A1'].font = Font(bold=True, size=15, color="1E3A8A", name="Arial")
    ws['A1'].fill = PatternFill("solid", fgColor="EFF6FF")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    STATUS_COLORS = {
        "قيد الدراسة":    ("FEF9C3", "854D0E"),
        "مقبول":          ("DCFCE7", "166534"),
        "مرفوض":          ("FEE2E2", "991B1B"),
        "تم الحل جزئياً": ("DBEAFE", "1E40AF"),
        "تم الحل":        ("D1FAE5", "065F46"),
    }

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill = PatternFill("solid", fgColor="1E3A8A")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[2].height = 28

    for ri, (_, row) in enumerate(df.iterrows(), 3):
        status_val = str(row.get("final_status", ""))
        bg_h, fg_h = STATUS_COLORS.get(status_val, ("F8FAFC", "1E293B"))
        for ci, col in enumerate(db_cols, 1):
            val = row.get(col, "")
            if pd.isna(val): val = ""
            c = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            c.fill = PatternFill("solid", fgColor=bg_h)
            c.font = Font(name="Arial", size=10, color=fg_h)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
        ws.row_dimensions[ri].height = 20

    COL_WIDTHS = {
        "complaint_number": 18, "submission_date": 14, "branch_name": 18,
        "complainant_name": 22, "complainant_role": 14, "client_card_number": 16,
        "client_phone": 14, "reception_method": 16, "complaint_summary": 35,
        "attached_documents": 25, "notification_date": 18,
        "response_method": 18, "investigation_summary": 35,
        "final_status": 16, "rejection_justification": 28, "created_by_name": 18
    }
    for ci, col in enumerate(db_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 16)

    ws.freeze_panes = "A3"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ===================================================================
# ===================== واجهات شكاوى العملاء =======================
# ===================================================================

def _parse_branches(user_branches):
    if isinstance(user_branches, str):
        try: return json.loads(user_branches)
        except: return [user_branches]
    return user_branches or []

def _status_badge(status):
    COLOR = {
        "قيد الدراسة":    ("#FEF9C3", "#92400E"),
        "مقبول":          ("#DCFCE7", "#166534"),
        "مرفوض":          ("#FEE2E2", "#991B1B"),
        "تم الحل جزئياً": ("#DBEAFE", "#1E40AF"),
        "تم الحل":        ("#D1FAE5", "#065F46"),
    }
    bg, fg = COLOR.get(str(status), ("#F1F5F9", "#475569"))
    return (
        f"<span style='background:{bg};color:{fg};font-weight:700;font-size:11px;"
        f"padding:3px 12px;border-radius:20px;white-space:nowrap;'>{status}</span>"
    )

def _request_status_badge(status):
    COLOR = {"pending": ("#FEF9C3","#92400E"), "approved": ("#DCFCE7","#166534"), "rejected": ("#FEE2E2","#991B1B")}
    LABEL = {"pending": "⏳ قيد الانتظار", "approved": "✅ موافق عليه", "rejected": "❌ مرفوض"}
    bg, fg = COLOR.get(status, ("#F1F5F9","#475569"))
    lbl = LABEL.get(status, status)
    return f"<span style='background:{bg};color:{fg};font-weight:700;font-size:11px;padding:3px 12px;border-radius:20px;'>{lbl}</span>"


# ── تبويب 1: إدخال شكوى ──────────────────────────────────────────
def show_entry_tab(user, is_admin, user_branches):
    st.markdown('<div class="sec-title">➕ تسجيل شكوى جديدة</div>', unsafe_allow_html=True)

    if is_admin:
        branch_options = [
            "فرع ابو المطامير","فرع اسيوط","فرع الاسكندرية",
            "فرع الاقصر","فرع المنيا","فرع دمياط","فرع سوهاج","فرع قنا"
        ]
    else:
        branch_options = user_branches

    if not branch_options:
        st.warning("⚠️ لا توجد فروع مخصصة لحسابك — تواصل مع الإدارة")
        return

    st.markdown(
        "<div style='background:linear-gradient(90deg,#eff6ff,#f0fdf4);"
        "border-radius:12px;padding:12px 18px;margin-bottom:18px;"
        "border:1px solid #bfdbfe;font-size:13px;color:#1e3a8a;'>"
        "📌 يُرجى ملء جميع الحقول المُشار إليها بـ <strong>(*)</strong> — "
        "سيتم توليد رقم قيد الشكوى تلقائياً بعد الحفظ."
        "</div>",
        unsafe_allow_html=True
    )

    with st.form("complaint_entry_form", clear_on_submit=True):
        st.markdown("#### 📋 بيانات مقدم الشكوى")
        c1, c2, c3 = st.columns(3)
        with c1:
            branch = st.selectbox("🏢 الفرع *", branch_options)
        with c2:
            sub_date = st.date_input("📅 تاريخ تقديم الشكوى *", value=datetime.now().date())
        with c3:
            reception = st.selectbox("📞 طريقة استقبال الشكوى *",
                ["تليفونيا", "وسائل تواصل اجتماعي", "بريد الكتروني", "حضور شخصي", "خطاب"])

        c4, c5, c6 = st.columns(3)
        with c4:
            comp_name = st.text_input("👤 اسم مقدم الشكوى *")
        with c5:
            comp_role = st.selectbox("🔖 صفة مقدم الشكوى",
                ["عميل", "ضامن", "عمل", "ذو صلة بعميل", "أخرى"])
        with c6:
            card_num = st.text_input("🪪 رقم بطاقة العميل")

        c7, c8 = st.columns(2)
        with c7:
            phone = st.text_input("📱 رقم هاتف العميل")
        with c8:
            docs = st.text_input("📎 بيان المستندات المرفقة (إن وُجدت)")

        st.markdown("#### 📝 تفاصيل الشكوى")
        summary = st.text_area("بيان موجز بموضوع الشكوى *", height=130,
            placeholder="اكتب هنا ملخصاً واضحاً لموضوع الشكوى...")

        submitted = st.form_submit_button("💾 تسجيل الشكوى", use_container_width=True)

        if submitted:
            errors = []
            if not comp_name.strip():  errors.append("اسم مقدم الشكوى")
            if not summary.strip():    errors.append("بيان موجز بموضوع الشكوى")
            if errors:
                st.error(f"❌ الحقول التالية مطلوبة: {' — '.join(errors)}")
            else:
                with st.spinner("⏳ جاري تسجيل الشكوى..."):
                    comp_num = generate_complaint_number()
                    data = {
                        "complaint_number":   comp_num,
                        "submission_date":    sub_date.isoformat(),
                        "branch_name":        branch,
                        "complainant_name":   comp_name.strip(),
                        "complainant_role":   comp_role,
                        "client_card_number": card_num.strip() or None,
                        "client_phone":       phone.strip() or None,
                        "reception_method":   reception,
                        "complaint_summary":  summary.strip(),
                        "attached_documents": docs.strip() or None,
                        "final_status":       "قيد الدراسة",
                        "created_by":         user.get('id'),
                        "created_by_name":    user.get('full_name'),
                    }
                    ok, result = insert_new_complaint(data, user)
                if ok:
                    st.success(f"✅ تم تسجيل الشكوى بنجاح!")
                    st.markdown(
                        f"<div style='background:#eff6ff;border:2px solid #2563eb;"
                        f"border-radius:12px;padding:16px;text-align:center;margin-top:10px;'>"
                        f"<div style='font-size:13px;color:#64748b;'>رقم قيد الشكوى</div>"
                        f"<div style='font-size:26px;font-weight:800;color:#1e3a8a;'>{comp_num}</div>"
                        f"<div style='font-size:12px;color:#64748b;margin-top:4px;'>"
                        f"احتفظ بهذا الرقم للمراجعة والمتابعة</div></div>",
                        unsafe_allow_html=True
                    )
                    st.balloons()
                else:
                    st.error(f"❌ خطأ في الحفظ: {result}")


# ── تبويب 2: قائمة الشكاوى ───────────────────────────────────────
def show_list_tab(user, is_admin, user_branches):
    st.markdown('<div class="sec-title">📋 قائمة الشكاوى المسجلة</div>', unsafe_allow_html=True)

    with st.spinner("جاري تحميل البيانات..."):
        df = fetch_all_complaints(is_admin, user_branches)

    if df.empty:
        st.info("📭 لا توجد شكاوى مسجلة حتى الآن")
        return

    st.markdown('<div class="filter-bar"><div class="filter-bar-title">🔍 أدوات البحث والتصفية</div>', unsafe_allow_html=True)
    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        br_opts = ["الكل"] + sorted(df['branch_name'].dropna().unique().tolist())
        sel_br = st.selectbox("🏢 الفرع", br_opts, key="cl_br")
    with fc2:
        st_opts = ["الكل"] + sorted(df['final_status'].dropna().unique().tolist())
        sel_st = st.selectbox("📊 موقف الشكوى", st_opts, key="cl_st")
    with fc3:
        rec_opts = ["الكل"] + sorted(df['reception_method'].dropna().unique().tolist())
        sel_rec = st.selectbox("📞 طريقة الاستقبال", rec_opts, key="cl_rec")
    with fc4:
        srch = st.text_input("🔎 بحث (اسم / رقم بطاقة / ملخص)", key="cl_srch")
    st.markdown('</div>', unsafe_allow_html=True)

    mask = pd.Series(True, index=df.index)
    if sel_br  != "الكل": mask &= df['branch_name']      == sel_br
    if sel_st  != "الكل": mask &= df['final_status']      == sel_st
    if sel_rec != "الكل": mask &= df['reception_method']  == sel_rec
    if srch:
        mask &= (
            df['complainant_name'].astype(str).str.contains(srch, case=False, na=False) |
            df['client_card_number'].astype(str).str.contains(srch, case=False, na=False) |
            df['complaint_summary'].astype(str).str.contains(srch, case=False, na=False) |
            df['complaint_number'].astype(str).str.contains(srch, case=False, na=False)
        )

    filtered = df[mask].reset_index(drop=True)
    if filtered.empty:
        st.warning("⚠️ لا توجد نتائج تطابق معايير البحث"); return

    cnt = filtered['final_status'].value_counts()
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-lbl">📊 إجمالي الشكاوى</div><div class="kpi-val">{len(filtered):,}</div></div>', unsafe_allow_html=True)
    with k2:
        v = cnt.get('قيد الدراسة', 0)
        st.markdown(f'<div class="kpi-card" style="border-top-color:#f59e0b"><div class="kpi-lbl">⏳ قيد الدراسة</div><div class="kpi-val" style="color:#d97706">{v:,}</div></div>', unsafe_allow_html=True)
    with k3:
        v = cnt.get('تم الحل', 0) + cnt.get('مقبول', 0) + cnt.get('تم الحل جزئياً', 0)
        st.markdown(f'<div class="kpi-card" style="border-top-color:#059669"><div class="kpi-lbl">✅ محلول / مقبول</div><div class="kpi-val" style="color:#059669">{v:,}</div></div>', unsafe_allow_html=True)
    with k4:
        v = cnt.get('مرفوض', 0)
        st.markdown(f'<div class="kpi-card" style="border-top-color:#dc2626"><div class="kpi-lbl">❌ مرفوض</div><div class="kpi-val" style="color:#dc2626">{v:,}</div></div>', unsafe_allow_html=True)

    tbl_html = """
    <div style='overflow-x:auto;margin-bottom:20px;'>
    <table style='width:100%;border-collapse:collapse;background:#fff;border-radius:14px;
                  overflow:hidden;box-shadow:0 4px 20px rgba(30,58,138,0.10);font-size:12px;'>
    <thead><tr style='background:#1e3a8a;'>
      <th style='padding:11px 12px;color:#fff;text-align:center;white-space:nowrap;'>رقم الشكوى</th>
      <th style='padding:11px 12px;color:#fff;text-align:center;white-space:nowrap;'>تاريخ التقديم</th>
      <th style='padding:11px 12px;color:#fff;text-align:center;white-space:nowrap;'>الفرع</th>
      <th style='padding:11px 12px;color:#fff;text-align:center;white-space:nowrap;'>اسم الشاكي</th>
      <th style='padding:11px 12px;color:#fff;text-align:center;white-space:nowrap;'>طريقة الاستقبال</th>
      <th style='padding:11px 12px;color:#fff;text-align:center;white-space:nowrap;'>موقف الشكوى</th>
      <th style='padding:11px 12px;color:#fff;text-align:center;white-space:nowrap;'>المدخل</th>
     </tr></thead><tbody>
    """
    for i, row in filtered.iterrows():
        bg = "#f8fafc" if i % 2 == 0 else "#ffffff"
        tbl_html += (
            f"<tr style='background:{bg};'>"
            f"<td style='padding:10px 12px;text-align:center;border-bottom:1px solid #e2e8f0;"
            f"font-weight:700;color:#1e3a8a;'>{row.get('complaint_number','—')}</td>"
            f"<td style='padding:10px 12px;text-align:center;border-bottom:1px solid #e2e8f0;'>"
            f"{str(row.get('submission_date','—'))[:10]}</td>"
            f"<td style='padding:10px 12px;text-align:center;border-bottom:1px solid #e2e8f0;'>"
            f"{row.get('branch_name','—')}</td>"
            f"<td style='padding:10px 12px;text-align:center;border-bottom:1px solid #e2e8f0;'>"
            f"{row.get('complainant_name','—')}</td>"
            f"<td style='padding:10px 12px;text-align:center;border-bottom:1px solid #e2e8f0;'>"
            f"{row.get('reception_method','—')}</td>"
            f"<td style='padding:10px 12px;text-align:center;border-bottom:1px solid #e2e8f0;'>"
            f"{_status_badge(row.get('final_status','—'))}</td>"
            f"<td style='padding:10px 12px;text-align:center;border-bottom:1px solid #e2e8f0;"
            f"font-size:11px;color:#64748b;'>{row.get('created_by_name','—')}</td>"
            f"</tr>"
        )
    tbl_html += "</tbody></table></div>"
    st.markdown(tbl_html, unsafe_allow_html=True)

    col_dl1, col_dl2 = st.columns([3, 1])
    with col_dl2:
        xls_data = generate_complaints_excel(filtered)
        st.download_button(
            "📥 تنزيل Excel",
            data=xls_data,
            file_name=f"شكاوى_العملاء_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.markdown('<div class="sec-title">🔍 عرض وتعديل شكوى</div>', unsafe_allow_html=True)

    opts = {
        f"{r['complaint_number']}  —  {r['complainant_name']}  ({r['branch_name']})": r
        for _, r in filtered.iterrows()
    }
    sel_lbl = st.selectbox("اختر الشكوى لعرض تفاصيلها أو طلب تعديل", list(opts.keys()), key="cl_sel")

    if not sel_lbl:
        return

    sel = opts[sel_lbl]

    with st.expander("📋 تفاصيل الشكوى الكاملة", expanded=True):
        d1, d2, d3 = st.columns(3)
        with d1:
            st.markdown(f"**رقم الشكوى:** {sel.get('complaint_number','—')}")
            st.markdown(f"**الفرع:** {sel.get('branch_name','—')}")
            st.markdown(f"**اسم مقدم الشكوى:** {sel.get('complainant_name','—')}")
            st.markdown(f"**صفة مقدم الشكوى:** {sel.get('complainant_role','—')}")
        with d2:
            st.markdown(f"**تاريخ التقديم:** {str(sel.get('submission_date','—'))[:10]}")
            st.markdown(f"**رقم بطاقة العميل:** {sel.get('client_card_number','—') or '—'}")
            st.markdown(f"**رقم الهاتف:** {sel.get('client_phone','—') or '—'}")
            st.markdown(f"**طريقة الاستقبال:** {sel.get('reception_method','—')}")
        with d3:
            st.markdown(f"**الموقف النهائي:** ", unsafe_allow_html=False)
            st.markdown(_status_badge(sel.get('final_status','—')), unsafe_allow_html=True)
            st.markdown(f"**تاريخ إبلاغ العميل:** {str(sel.get('notification_date','—') or '—')[:10]}")
            st.markdown(f"**طريقة الرد:** {sel.get('response_method','—') or '—'}")
            st.markdown(f"**المدخل:** {sel.get('created_by_name','—')}")

        st.markdown("**📝 موجز الشكوى:**")
        st.info(sel.get('complaint_summary', '—'))

        if sel.get('attached_documents'):
            st.markdown(f"**📎 المستندات المرفقة:** {sel['attached_documents']}")

        if sel.get('investigation_summary'):
            st.markdown("**🔍 موجز نتيجة الفحص:**")
            st.success(sel['investigation_summary'])

        if sel.get('rejection_justification'):
            st.markdown("**⚠️ مبررات الرفض:**")
            st.error(sel['rejection_justification'])

    with st.expander("✏️ طلب تعديل هذه الشكوى", expanded=False):
        st.info(
            "💡 يمكنك تعديل الحقول التالية — سيُرسل الطلب للإدارة وسيتم إشعارك بالنتيجة."
        )

        with st.form(f"edit_form_{sel['id'][:8]}"):
            RESP_OPTS = ["تليفونيا", "وسائل تواصل اجتماعي", "بريد الكتروني", "خطاب"]
            STATUS_OPTS = ["قيد الدراسة", "مقبول", "مرفوض", "تم الحل جزئياً", "تم الحل"]

            # عرض الحقول التي يمكن تعديلها للمستخدم العادي
            if is_admin:
                # المسؤول يمكنه تعديل كل شيء
                ec1, ec2 = st.columns(2)
                with ec1:
                    curr_notif = sel.get('notification_date')
                    if curr_notif is None or pd.isna(curr_notif):
                        notif_val = datetime.now().date()
                    else:
                        try:
                            notif_val = pd.to_datetime(curr_notif).date()
                        except:
                            notif_val = datetime.now().date()
                    new_notif = st.date_input("📅 تاريخ إبلاغ العميل بنتيجة الشكوى", value=notif_val)
                with ec2:
                    curr_resp = sel.get('response_method') or RESP_OPTS[0]
                    resp_idx = RESP_OPTS.index(curr_resp) if curr_resp in RESP_OPTS else 0
                    new_resp = st.selectbox("📞 طريقة الرد على الشاكي", RESP_OPTS, index=resp_idx)

                curr_status = sel.get('final_status') or STATUS_OPTS[0]
                st_idx = STATUS_OPTS.index(curr_status) if curr_status in STATUS_OPTS else 0
                new_status = st.selectbox("📊 موقف الشكوى النهائي", STATUS_OPTS, index=st_idx)

                new_inv = st.text_area(
                    "🔍 موجز ما انتهى إليه فحص الشكوى من رأى",
                    value=sel.get('investigation_summary') or "", height=90
                )

                new_rej = ""
                if new_status == "مرفوض":
                    new_rej = st.text_area(
                        "⚠️ المبررات في حالة رفض الشكوى",
                        value=sel.get('rejection_justification') or "", height=70
                    )
            else:
                # المستخدم العادي: يعرض فقط الحقول القابلة للتعديل (بدون موقف الشكوى)
                st.info("📝 يمكنك اقتراح تغييرات على الشكوى. سيتم مراجعتها من قبل الإدارة.")
                
                # حقل تاريخ الإبلاغ
                curr_notif = sel.get('notification_date')
                if curr_notif is None or pd.isna(curr_notif):
                    notif_val = datetime.now().date()
                else:
                    try:
                        notif_val = pd.to_datetime(curr_notif).date()
                    except:
                        notif_val = datetime.now().date()
                new_notif = st.date_input("📅 تاريخ إبلاغ العميل بنتيجة الشكوى", value=notif_val)
                
                # حقل طريقة الرد
                curr_resp = sel.get('response_method') or RESP_OPTS[0]
                resp_idx = RESP_OPTS.index(curr_resp) if curr_resp in RESP_OPTS else 0
                new_resp = st.selectbox("📞 طريقة الرد على الشاكي", RESP_OPTS, index=resp_idx)
                
                # حقل موجز الفحص
                new_inv = st.text_area(
                    "🔍 موجز ما انتهى إليه فحص الشكوى من رأى",
                    value=sel.get('investigation_summary') or "", height=90
                )
                
                # المستخدم العادي لا يمكنه تغيير حالة الشكوى
                new_status = sel.get('final_status') or STATUS_OPTS[0]
                st.info(f"⚠️ حالة الشكوى الحالية: **{new_status}** (لا يمكن تغييرها إلا من قبل الإدارة)")
                
                new_rej = ""
                if new_status == "مرفوض" and sel.get('rejection_justification'):
                    st.info(f"سبب الرفض الحالي: {sel.get('rejection_justification')}")

            edit_note_txt = st.text_input("💬 ملاحظة إضافية لطلب التعديل (اختياري)")

            btn_submit = st.form_submit_button("📤 إرسال طلب التعديل", use_container_width=True)

            if btn_submit:
                changes = {}
                
                # مقارنة القيم فقط إذا كانت مختلفة
                old_n = str(sel.get('notification_date') or "")
                new_n = new_notif.isoformat()
                if old_n != new_n and old_n != "None" and old_n != "":
                    changes['notification_date'] = {"old": old_n if old_n != "None" else "", "new": new_n, "label": "تاريخ إبلاغ العميل"}

                old_r = sel.get('response_method') or ""
                if old_r != new_resp:
                    changes['response_method'] = {"old": old_r if old_r != "None" else "", "new": new_resp, "label": "طريقة الرد على الشاكي"}

                old_i = sel.get('investigation_summary') or ""
                if old_i != new_inv.strip() and new_inv.strip() != "":
                    changes['investigation_summary'] = {"old": old_i, "new": new_inv.strip() or "", "label": "موجز نتيجة الفحص"}

                # فقط المسؤول يمكنه تغيير الحالة
                if is_admin:
                    old_s = sel.get('final_status') or ""
                    if old_s != new_status:
                        changes['final_status'] = {"old": old_s, "new": new_status, "label": "موقف الشكوى النهائي"}

                    if new_status == "مرفوض":
                        old_rej = sel.get('rejection_justification') or ""
                        if old_rej != new_rej.strip():
                            changes['rejection_justification'] = {"old": old_rej, "new": new_rej.strip() or "", "label": "مبررات الرفض"}

                if not changes:
                    st.warning("⚠️ لم تقم بإجراء أي تغييرات على الحقول! قم بتعديل قيمة أحد الحقول ثم أرسل الطلب.")
                else:
                    if edit_note_txt:
                        changes['_note'] = edit_note_txt
                    with st.spinner("جاري إرسال طلب التعديل..."):
                        ok, result = submit_edit_request_db(
                            sel['id'], sel['complaint_number'],
                            sel['branch_name'], changes, user
                        )
                    if ok:
                        st.success("✅ تم إرسال طلب التعديل — في انتظار موافقة الإدارة.")
                        st.rerun()
                    else:
                        st.error(f"❌ خطأ: {result}")

# ── تبويب 3: طلبات التعديل ───────────────────────────────────────
def show_edit_requests_tab(user, is_admin):
    user_id = user.get('id')
    title = "📋 طلبات التعديل المعلّقة" if is_admin else "📋 طلباتي للتعديل"
    st.markdown(f'<div class="sec-title">{title}</div>', unsafe_allow_html=True)

    status_filter_opts = ["الكل", "pending", "approved", "rejected"]
    STATUS_AR = {"الكل": "الكل", "pending": "⏳ قيد الانتظار", "approved": "✅ موافق", "rejected": "❌ مرفوض"}

    fc1, fc2 = st.columns([2, 4])
    with fc1:
        sel_sf = st.selectbox("تصفية بحسب الحالة",
            [STATUS_AR[s] for s in status_filter_opts], key="er_sf")
        raw_sf = {v: k for k, v in STATUS_AR.items()}.get(sel_sf, "الكل")
        sf = None if raw_sf == "الكل" else raw_sf

    with st.spinner("جاري التحميل..."):
        df_req = fetch_edit_requests_db(is_admin=is_admin, user_id=user_id, status_filter=sf)

    if df_req.empty:
        st.info("📭 لا توجد طلبات تعديل حالياً")
        return

    for _, req in df_req.iterrows():
        changes = req.get('changes', {})
        if isinstance(changes, str):
            try: changes = json.loads(changes)
            except: changes = {}

        req_status = req.get('status', 'pending')
        note_field = changes.pop('_note', None)

        border_color = {"pending": "#f59e0b", "approved": "#059669", "rejected": "#dc2626"}.get(req_status, "#2563eb")
        bg_color     = {"pending": "#fffbeb", "approved": "#f0fdf4", "rejected": "#fef2f2"}.get(req_status, "#f8fafc")

        st.markdown(
            f"<div style='background:{bg_color};border:1.5px solid {border_color};"
            f"border-right:5px solid {border_color};border-radius:14px;padding:16px 20px;"
            f"margin-bottom:14px;'>",
            unsafe_allow_html=True
        )

        hd1, hd2, hd3 = st.columns([3, 2, 2])
        with hd1:
            st.markdown(f"**📝 {req.get('complaint_number','—')}** — فرع {req.get('branch_name','—')}")
        with hd2:
            st.markdown(f"**الطالب:** {req.get('requester_name','—')}")
        with hd3:
            st.markdown(_request_status_badge(req_status), unsafe_allow_html=True)

        req_time = str(req.get('requested_at',''))[:16].replace('T',' ')
        st.markdown(f"<span style='font-size:11px;color:#64748b;'>🕐 {req_time}</span>", unsafe_allow_html=True)

        if changes:
            chg_rows = ""
            for field, vals in changes.items():
                if field == "_note": continue
                label = vals.get('label', field)
                old_v = vals.get('old', '—') or '—'
                new_v = vals.get('new', '—') or '—'
                chg_rows += (
                    f"<tr>"
                    f"<td style='padding:8px 12px;border-bottom:1px solid #e2e8f0;font-weight:600;color:#475569;'>{label}</td>"
                    f"<td style='padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#dc2626;text-decoration:line-through;'>{old_v}</td>"
                    f"<td style='padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#059669;font-weight:700;'>{new_v}</td>"
                    f"</tr>"
                )
            st.markdown(
                f"<table style='width:100%;border-collapse:collapse;margin:10px 0;"
                f"border-radius:10px;overflow:hidden;font-size:12px;'>"
                f"<thead><tr style='background:#1e3a8a;'>"
                f"<th style='padding:8px 12px;color:#fff;text-align:right;'>الحقل</th>"
                f"<th style='padding:8px 12px;color:#fca5a5;text-align:center;'>القيمة القديمة</th>"
                f"<th style='padding:8px 12px;color:#6ee7b7;text-align:center;'>القيمة الجديدة</th>"
                f"</tr></thead><tbody>{chg_rows}</tbody></table>",
                unsafe_allow_html=True
            )

        if note_field:
            st.markdown(f"💬 **ملاحظة الطالب:** {note_field}")

        if req_status != "pending" and req.get('admin_note'):
            st.markdown(f"📌 **ملاحظة الإدارة:** {req['admin_note']}")

        if req.get('reviewed_at'):
            rev_time = str(req.get('reviewed_at',''))[:16].replace('T',' ')
            st.markdown(f"<span style='font-size:11px;color:#64748b;'>🔍 راجعه {req.get('reviewed_by_name','—')} في {rev_time}</span>", unsafe_allow_html=True)

        if is_admin and req_status == "pending":
            req_id = req.get('id')
            comp_id = req.get('complaint_id')
            requester_id = req.get('requested_by')

            with st.form(f"admin_form_{req_id[:8]}"):
                admin_note_txt = st.text_input("💬 ملاحظة (اختياري)", key=f"an_{req_id[:8]}")
                ba, bb = st.columns(2)
                with ba:
                    approve_btn = st.form_submit_button("✅ موافقة", use_container_width=True)
                with bb:
                    reject_btn = st.form_submit_button("❌ رفض", use_container_width=True)

                if approve_btn:
                    with st.spinner("جاري تطبيق التعديل..."):
                        ok = process_edit_request_db(
                            req_id, comp_id, changes, "approved",
                            user, admin_note_txt, requester_id
                        )
                    if ok: st.success("✅ تمت الموافقة وتطبيق التعديل!"); st.rerun()
                    else:  st.error("❌ حدث خطأ!")

                if reject_btn:
                    with st.spinner("جاري الرفض..."):
                        ok = process_edit_request_db(
                            req_id, comp_id, changes, "rejected",
                            user, admin_note_txt, requester_id
                        )
                    if ok: st.success("تم رفض الطلب وإشعار المستخدم."); st.rerun()
                    else:  st.error("❌ حدث خطأ!")

        st.markdown("</div>", unsafe_allow_html=True)


# ── تبويب 4: الإشعارات ───────────────────────────────────────────
def show_notifications_tab(user_id):
    st.markdown('<div class="sec-title">🔔 إشعاراتي</div>', unsafe_allow_html=True)

    with st.spinner("جاري تحميل الإشعارات..."):
        df_notif = fetch_notifications_db(user_id)

    if df_notif.empty:
        st.info("📭 لا توجد إشعارات حتى الآن")
        return

    unread = df_notif[df_notif['is_read'] == False]
    if not unread.empty:
        col_mark, _ = st.columns([2, 4])
        with col_mark:
            if st.button("✅ تحديد الكل كمقروء", key="mark_read"):
                mark_all_notifications_read(user_id)
                st.success("تم تحديد كل الإشعارات كمقروءة")
                st.rerun()
        st.markdown(
            f"<div style='background:#eff6ff;border-radius:10px;padding:10px 16px;"
            f"margin-bottom:12px;font-size:13px;color:#1e3a8a;font-weight:600;'>"
            f"🔔 لديك {len(unread)} إشعار غير مقروء</div>",
            unsafe_allow_html=True
        )

    ICON_MAP = {
        "new_complaint": ("📝", "notif-card"),
        "edit_request":  ("✏️", "notif-card edit_request"),
        "approved":      ("✅", "notif-card approved"),
        "rejected":      ("❌", "notif-card rejected"),
    }

    for _, n in df_notif.iterrows():
        ntype = n.get('notification_type', '')
        is_unread = not n.get('is_read', True)
        icon, card_cls = ICON_MAP.get(ntype, ("🔔", "notif-card"))
        if is_unread:
            card_cls += " unread"

        ts = str(n.get('created_at', ''))[:16].replace('T', ' ')
        badge = "<span class='notif-badge'>جديد</span>" if is_unread else ""

        st.markdown(
            f"<div class='{card_cls}'>"
            f"{badge}"
            f"<div class='notif-title'>{icon} {n.get('title','')}</div>"
            f"<div class='notif-msg'>{n.get('message','')}</div>"
            f"<div class='notif-time'>🕐 {ts}</div>"
            f"</div>",
            unsafe_allow_html=True
        )


# ── الصفحة الرئيسية للشكاوى ──────────────────────────────────────
def complaints_page():
    if 'user' not in st.session_state:
        st.warning("⚠️ يجب تسجيل الدخول أولاً"); st.stop()

    user = st.session_state['user']
    show_header(user)

    col_back, col_title = st.columns([1, 6])
    with col_back:
        if st.button("🏠 الرئيسية"):
            st.query_params.clear(); st.rerun()
    with col_title:
        st.markdown('<div class="sec-title">📝 سجل شكاوى العملاء</div>', unsafe_allow_html=True)

    is_admin     = user.get('role') == 'admin'
    user_branches = _parse_branches(user.get('branches', []))
    user_id      = user.get('id')

    unread = count_unread_notifications(user_id)
    notif_lbl = f"🔔 الإشعارات  ({unread})" if unread > 0 else "🔔 الإشعارات"

    if unread > 0:
        st.markdown(
            f"<div style='background:linear-gradient(90deg,#fef2f2,#fff7ed);"
            f"border:1.5px solid #ef4444;border-radius:12px;padding:10px 18px;"
            f"margin-bottom:14px;display:flex;align-items:center;gap:10px;'>"
            f"<span style='font-size:20px;'>🔔</span>"
            f"<span style='font-weight:700;color:#991b1b;font-size:13px;'>"
            f"لديك {unread} إشعار غير مقروء — تحقق من تبويب الإشعارات</span>"
            f"</div>",
            unsafe_allow_html=True
        )

    tab1, tab2, tab3, tab4 = st.tabs([
        "➕ إدخال شكوى جديدة",
        "📋 قائمة الشكاوى",
        "✏️ طلبات التعديل",
        notif_lbl
    ])

    with tab1:
        show_entry_tab(user, is_admin, user_branches)
    with tab2:
        show_list_tab(user, is_admin, user_branches)
    with tab3:
        show_edit_requests_tab(user, is_admin)
    with tab4:
        show_notifications_tab(user_id)


# ===================================================================
# ===================== صفحة فوري =====================
# ===================================================================

def reports_page():
    if 'user' not in st.session_state:
        st.warning("⚠️ يجب تسجيل الدخول أولاً"); st.stop()

    user=st.session_state['user']
    show_header(user)

    col_back, col_title = st.columns([1,6])
    with col_back:
        if st.button("🏠 الرئيسية"):
            st.query_params.clear(); st.rerun()
    with col_title:
        st.markdown('<div class="sec-title">💳 سداد فوري & Opay</div>', unsafe_allow_html=True)

    is_admin      = user.get('role') == 'admin'
    user_branches = user.get('branches', [])

    with st.spinner("جاري تحميل البيانات..."):
        df_raw = fetch_reports_data()
    if df_raw.empty:
        st.info("📭 لا توجد بيانات متاحة"); return

    df_acc  = df_raw if is_admin else df_raw[df_raw['branch_name'].isin(user_branches)]
    v_dates = df_acc['تاريخ الدفع'].dropna()
    if v_dates.empty:
        st.info("📭 لا توجد تواريخ متاحة"); return

    st.markdown('<div class="filter-bar"><div class="filter-bar-title">🔍 أدوات البحث والتصفية</div>', unsafe_allow_html=True)

    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1: start_d = st.date_input("📅 من تاريخ", v_dates.min().date(), key="r_from")
    with fc2: end_d   = st.date_input("📅 إلى تاريخ", v_dates.max().date(), key="r_to")
    with fc3:
        all_branches_r = ["الكل"] + sorted(df_acc['branch_name'].dropna().unique().tolist())
        sel_branch = st.selectbox("🏢 الفرع", all_branches_r, key="r_branch")
    with fc4:
        codes = ["الكل"] + sorted(df_acc['كود الخدمة'].dropna().unique().tolist())
        sel_code = st.selectbox("🏷️ كود الخدمة", codes, key="r_code")

    fc5, fc6 = st.columns(2)
    with fc5: s_name = st.text_input("🔎 بحث بالاسم أو كود العميل", key="r_search")
    with fc6: s_nat  = st.text_input("🆔 بحث بالرقم القومي / رقم المرجع", key="r_nat")
    st.markdown('</div>', unsafe_allow_html=True)

    mask = (df_acc['تاريخ الدفع'].dt.date >= start_d) & (df_acc['تاريخ الدفع'].dt.date <= end_d)
    if sel_branch != "الكل": mask &= (df_acc['branch_name'] == sel_branch)
    if sel_code   != "الكل": mask &= (df_acc['كود الخدمة'] == sel_code)
    if s_name:
        mask &= (df_acc['client_name'].astype(str).str.contains(s_name, na=False, case=False) |
                 df_acc['client_code'].astype(str).str.contains(s_name, na=False, case=False))
    if s_nat:
        mask &= df_acc['رقم المرجع'].astype(str).str.contains(s_nat, na=False, case=False)

    final_df = df_acc.loc[mask]
    if final_df.empty:
        st.warning("⚠️ لا توجد بيانات تطابق معايير البحث"); return

    code_stats = (
        final_df.groupby('كود الخدمة')
        .agg(عدد=('المبلغ', 'count'), مبلغ=('المبلغ', 'sum'))
        .reset_index()
        .sort_values('عدد', ascending=False)
    )

    k1, k2 = st.columns(2)
    with k1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-lbl">💰 إجمالي المبالغ</div><div class="kpi-val">{final_df["المبلغ"].sum():,.0f} ج.م</div></div>', unsafe_allow_html=True)
    with k2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-lbl">📊 عدد العمليات</div><div class="kpi-val">{len(final_df):,} حركة</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="sec-title">📋 تفاصيل الأكواد</div>', unsafe_allow_html=True)
    num_codes = len(code_stats)
    cols_count = min(num_codes, 4) if num_codes > 0 else 1
    code_cols = st.columns(cols_count)
    for i, row in enumerate(code_stats.itertuples(index=False)):
        col_idx = i % cols_count
        with code_cols[col_idx]:
            st.markdown(
                f'<div class="code-card"><div class="code-card-name">🏷️ {row[0]}</div>'
                f'<div class="code-card-count">{int(row[1]):,}</div><div class="code-card-unit">عملية</div>'
                f'<div class="code-card-amt">{row[2]:,.0f} ج.م</div><div class="code-card-amt-l">إجمالي المبلغ</div></div>',
                unsafe_allow_html=True
            )

    with st.expander("📅 الملخص اليومي", expanded=False):
        CODE_COLORS_HTML = ["#dbeafe","#dcfce7","#fef9c3","#fce7f3","#ede9fe","#ffedd5","#cffafe","#fee2e2"]
        CODE_TEXT_HTML   = ["#1e40af","#166534","#854d0e","#9d174d","#6b21a8","#9a3412","#0e7490","#991b1b"]
        all_codes = sorted(final_df['كود الخدمة'].dropna().unique().tolist())
        code_color_map_html = {c: (CODE_COLORS_HTML[i%len(CODE_COLORS_HTML)],CODE_TEXT_HTML[i%len(CODE_TEXT_HTML)]) for i,c in enumerate(all_codes)}
        temp_daily = final_df.copy(); temp_daily['_date'] = temp_daily['تاريخ الدفع'].dt.date
        dates_sorted = sorted(temp_daily['_date'].dropna().unique())
        rows_html = ""
        for d in dates_sorted:
            df_day = temp_daily[temp_daily['_date']==d]
            day_total_count=len(df_day); day_total_amt=df_day['المبلغ'].sum()
            first_code=True
            codes_in_day=[c for c in all_codes if not df_day[df_day['كود الخدمة']==c].empty]
            rowspan_val=len(codes_in_day)
            for code in all_codes:
                df_code=df_day[df_day['كود الخدمة']==code]
                if df_code.empty: continue
                bg,fg=code_color_map_html[code]; cnt=len(df_code); amt=df_code['المبلغ'].sum()
                date_cell=(f"<td rowspan='{rowspan_val}' style='padding:10px 14px;text-align:center;border-bottom:2px solid #cbd5e1;border-left:1px solid #e2e8f0;font-weight:700;color:#1e3a8a;vertical-align:middle;background:#f8fafc;font-size:13px;white-space:nowrap;'>{d}</td>") if first_code else ""
                rows_html+=(f"<tr>"+date_cell+f"<td style='padding:9px 14px;text-align:center;border-bottom:1px solid #e2e8f0;background:{bg};color:{fg};font-weight:700;font-size:12px;'>{code}</td><td style='padding:9px 14px;text-align:center;border-bottom:1px solid #e2e8f0;background:{bg};color:{fg};font-weight:700;'>{cnt:,}</td><td style='padding:9px 14px;text-align:center;border-bottom:1px solid #e2e8f0;background:{bg};color:{fg};font-weight:700;'>{amt:,.2f}</td></tr>"); first_code=False
            rows_html+=(f"<tr style='background:#1e3a8a;'><td style='padding:10px 14px;text-align:center;color:#fff;font-weight:800;border-bottom:2px solid #60a5fa;font-size:12px;'>✦ إجمالي {d}</td><td style='padding:10px;text-align:center;color:#bfdbfe;font-weight:800;border-bottom:2px solid #60a5fa;'>الكل</td><td style='padding:10px;text-align:center;color:#fff;font-weight:800;border-bottom:2px solid #60a5fa;'>{day_total_count:,}</td><td style='padding:10px;text-align:center;color:#fde68a;font-weight:800;border-bottom:2px solid #60a5fa;'>{day_total_amt:,.2f}</td></tr>")
        grand_count=len(final_df); grand_amt=final_df['المبلغ'].sum()
        rows_html+=(f"<tr style='background:#0f172a;'><td style='padding:12px 14px;text-align:center;color:#fff;font-weight:800;'>الإجمالي الكلي</td><td style='padding:12px;text-align:center;color:#93c5fd;font-weight:800;'>—</td><td style='padding:12px;text-align:center;color:#fff;font-weight:800;'>{grand_count:,}</td><td style='padding:12px;text-align:center;color:#fde68a;font-weight:800;'>{grand_amt:,.2f}</td></tr>")
        th="padding:12px 14px;text-align:center;color:#fff;font-size:13px;font-weight:700;white-space:nowrap;"
        legend_html=" &nbsp; ".join([f"<span style='background:{code_color_map_html[c][0]};color:{code_color_map_html[c][1]};padding:3px 10px;border-radius:10px;font-size:12px;font-weight:700;'>{c}</span>" for c in all_codes])
        st.markdown(f"<div style='margin-bottom:10px;'>🎨 <strong>دليل الألوان:</strong> &nbsp; {legend_html}</div>", unsafe_allow_html=True)
        st.markdown(f"<div style='overflow-x:auto;'><table style='width:100%;border-collapse:collapse;background:#fff;border-radius:14px;overflow:hidden;box-shadow:0 4px 20px rgba(30,58,138,0.12);font-size:13px;'><thead><tr style='background:#1e3a8a;'><th style='{th}'>التاريخ</th><th style='{th}'>كود الخدمة</th><th style='{th}'>عدد الحركات</th><th style='{th}'>إجمالي المبلغ (ج.م)</th></tr></thead><tbody>{rows_html}</tbody></table></div>", unsafe_allow_html=True)

    with st.expander("🏢 ملخص الفروع اليومي", expanded=False):
        _CODE_COLORS_HTML=["#dbeafe","#dcfce7","#fef9c3","#fce7f3","#ede9fe","#ffedd5","#cffafe","#fee2e2"]
        _CODE_TEXT_HTML=["#1e40af","#166534","#854d0e","#9d174d","#6b21a8","#9a3412","#0e7490","#991b1b"]
        _all_codes=sorted(final_df['كود الخدمة'].dropna().unique().tolist())
        _code_color_map={c:(_CODE_COLORS_HTML[i%len(_CODE_COLORS_HTML)],_CODE_TEXT_HTML[i%len(_CODE_TEXT_HTML)]) for i,c in enumerate(_all_codes)}
        _temp=final_df.copy(); _temp['_date']=_temp['تاريخ الدفع'].dt.date
        _all_branches=sorted(_temp['branch_name'].dropna().unique().tolist())
        _dates_sorted=sorted(_temp['_date'].dropna().unique())
        _legend=" &nbsp; ".join([f"<span style='background:{_code_color_map[c][0]};color:{_code_color_map[c][1]};padding:3px 10px;border-radius:10px;font-size:12px;font-weight:700;'>{c}</span>" for c in _all_codes])
        st.markdown(f"<div style='margin-bottom:10px;'>🎨 <strong>دليل الألوان:</strong> &nbsp; {_legend}</div>", unsafe_allow_html=True)
        _th="padding:11px 10px;text-align:center;color:#fff;font-size:12px;font-weight:700;white-space:nowrap;"
        _th_r="padding:11px 12px;text-align:right;color:#fff;font-size:12px;font-weight:700;white-space:nowrap;"
        _rows=""
        for br in _all_branches:
            df_br=_temp[_temp['branch_name']==br]
            if df_br.empty: continue
            br_day_code_rows=sum(sum(1 for c in _all_codes if not df_br[(df_br['_date']==d)&(df_br['كود الخدمة']==c)].empty) for d in _dates_sorted if not df_br[df_br['_date']==d].empty)
            br_day_total_rows=sum(1 for d in _dates_sorted if not df_br[df_br['_date']==d].empty)
            br_rowspan=br_day_code_rows+br_day_total_rows
            first_day_in_br=True
            for d in _dates_sorted:
                df_day_br=df_br[df_br['_date']==d]
                if df_day_br.empty: continue
                day_cnt=len(df_day_br); day_amt=df_day_br['المبلغ'].sum()
                codes_here=[c for c in _all_codes if not df_day_br[df_day_br['كود الخدمة']==c].empty]
                day_rowspan=len(codes_here)+1; first_code_in_day=True
                for code in _all_codes:
                    df_dc=df_day_br[df_day_br['كود الخدمة']==code]
                    if df_dc.empty: continue
                    bg,fg=_code_color_map[code]; cnt=len(df_dc); amt=df_dc['المبلغ'].sum()
                    br_cell=(f"<td rowspan='{br_rowspan}' style='padding:10px 12px;text-align:center;border-bottom:2px solid #1e3a8a;border-left:1px solid #e2e8f0;font-weight:800;color:#fff;vertical-align:middle;background:linear-gradient(180deg,#1e3a8a,#2563eb);font-size:12px;writing-mode:vertical-rl;'>{br}</td>") if (first_day_in_br and first_code_in_day) else ""
                    day_cell=(f"<td rowspan='{day_rowspan}' style='padding:9px 10px;text-align:center;border-bottom:1px solid #cbd5e1;font-weight:700;color:#1e3a8a;vertical-align:middle;background:#f8fafc;font-size:12px;white-space:nowrap;'>{d}</td>") if first_code_in_day else ""
                    _rows+=(f"<tr>"+br_cell+day_cell+f"<td style='padding:8px 10px;text-align:center;border-bottom:1px solid #e2e8f0;background:{bg};color:{fg};font-weight:700;font-size:11px;'>{code}</td><td style='padding:8px 10px;text-align:center;border-bottom:1px solid #e2e8f0;background:{bg};color:{fg};font-weight:700;'>{cnt:,}</td><td style='padding:8px 10px;text-align:center;border-bottom:1px solid #e2e8f0;background:{bg};color:{fg};font-weight:700;'>{amt:,.2f}</td></tr>")
                    first_code_in_day=False; first_day_in_br=False
                _rows+=(f"<tr style='background:#0f172a;'><td style='padding:9px 10px;text-align:center;color:#93c5fd;font-weight:800;border-bottom:1.5px solid #334155;font-size:11px;'>✦ {d}</td><td style='padding:9px 10px;text-align:center;color:#bfdbfe;font-weight:800;border-bottom:1.5px solid #334155;font-size:11px;'>الكل</td><td style='padding:9px 10px;text-align:center;color:#fff;font-weight:800;border-bottom:1.5px solid #334155;'>{day_cnt:,}</td><td style='padding:9px 10px;text-align:center;color:#fde68a;font-weight:800;border-bottom:1.5px solid #334155;'>{day_amt:,.2f}</td></tr>")
        _rows+=(f"<tr style='background:#1e3a8a;'><td colspan='2' style='padding:12px;text-align:center;color:#fff;font-weight:800;font-size:13px;'>الإجمالي الكلي</td><td style='padding:12px;text-align:center;color:#bfdbfe;font-weight:800;'>—</td><td style='padding:12px;text-align:center;color:#fff;font-weight:800;'>{len(final_df):,}</td><td style='padding:12px;text-align:center;color:#fde68a;font-weight:800;'>{final_df['المبلغ'].sum():,.2f}</td></tr>")
        st.markdown(f"<div style='overflow-x:auto;'><table style='width:100%;border-collapse:collapse;background:#fff;border-radius:14px;overflow:hidden;box-shadow:0 4px 20px rgba(30,58,138,0.12);font-size:12px;'><thead><tr style='background:#1e3a8a;'><th style='{_th_r}'>الفرع</th><th style='{_th}'>التاريخ</th><th style='{_th}'>كود الخدمة</th><th style='{_th}'>عدد الحركات</th><th style='{_th}'>إجمالي المبلغ (ج.م)</th><tr></thead><tbody>{_rows}</tbody></table></div>", unsafe_allow_html=True)

    st.markdown('<div class="sec-title">📋 البيانات التفصيلية</div>', unsafe_allow_html=True)
    display_df = final_df.copy().rename(columns={'client_code':'كود العميل','client_name':'اسم العميل','branch_name':'الفرع'})
    drop_cols = [c for c in display_df.columns if c.startswith('_') or c == 'id']
    display_df = display_df.drop(columns=drop_cols, errors='ignore')
    if 'تاريخ الدفع' in display_df.columns:
        display_df['تاريخ الدفع'] = display_df['تاريخ الدفع'].dt.strftime('%Y-%m-%d')
    show_cols = ['الفرع','كود العميل','اسم العميل','رقم المرجع','تاريخ الدفع','وقت الدفع','المبلغ','كود الخدمة']
    show_cols = [c for c in show_cols if c in display_df.columns]
    view_df   = display_df[show_cols]
    col_config = {
        'الفرع':st.column_config.TextColumn('الفرع',width=110),'كود العميل':st.column_config.TextColumn('كود العميل',width=85),
        'اسم العميل':st.column_config.TextColumn('اسم العميل',width=130),'رقم المرجع':st.column_config.TextColumn('رقم المرجع',width=120),
        'تاريخ الدفع':st.column_config.TextColumn('تاريخ الدفع',width=90),'وقت الدفع':st.column_config.TextColumn('وقت الدفع',width=75),
        'المبلغ':st.column_config.NumberColumn('المبلغ',width=85,format='%.2f'),'كود الخدمة':st.column_config.TextColumn('كود الخدمة',width=80),
    }
    st.dataframe(view_df, use_container_width=True, hide_index=True, height=420, column_config=col_config)

    st.markdown('<div class="sec-title">📥 تحميل التقرير</div>', unsafe_allow_html=True)
    dl1,dl2,dl3=st.columns([2,2,2])
    with dl1:
        split_mode=st.radio("نوع التنزيل",["كل البيانات في شيت واحد","تقسيم يوم يوم (شيت لكل يوم)"],key="r_split")
    with dl2:
        selected_day="كل الأيام"
        if split_mode=="تقسيم يوم يوم (شيت لكل يوم)":
            av=sorted(final_df['تاريخ الدفع'].dropna().dt.date.unique())
            selected_day=st.selectbox("اختر اليوم",["كل الأيام"]+[str(d) for d in av],key="r_day")
    with dl3:
        if split_mode=="تقسيم يوم يوم (شيت لكل يوم)":
            if selected_day=="كل الأيام":
                xls=generate_reports_excel_daily(display_df,final_df); fname=f"تقرير_كل_الأيام_{datetime.now().date()}.xlsx"
            else:
                sd=pd.to_datetime(selected_day).date()
                ddf=display_df[pd.to_datetime(display_df['تاريخ الدفع'],errors='coerce').dt.date==sd]
                xls=generate_reports_excel_single(ddf,sheet_title=selected_day,report_title=f"تقرير سدادات يوم {selected_day}"); fname=f"تقرير_{selected_day}.xlsx"
        else:
            xls=generate_reports_excel_single(display_df,report_title=f"تقرير السدادات - {start_d} إلى {end_d}"); fname=f"تقرير_{datetime.now().date()}.xlsx"
        st.download_button("📊 تحميل Excel",data=xls,file_name=fname,mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)


# ===================================================================
# ===================== صفحة الأقساط المستحقة =====================
# ===================================================================

def outstanding_page():
    if 'user' not in st.session_state:
        st.warning("⚠️ يجب تسجيل الدخول أولاً"); st.stop()

    user=st.session_state['user']
    show_header(user)

    col_back,col_title=st.columns([1,6])
    with col_back:
        if st.button("🏠 الرئيسية"):
            st.query_params.clear(); st.rerun()
    with col_title:
        st.markdown('<div class="sec-title">📋 الأقساط المستحقة — فوري & Opay</div>',unsafe_allow_html=True)

    is_admin      = user.get('role')=='admin'
    branches_list = user.get('branches',[])

    if isinstance(branches_list,str):
        try:    branches_list=json.loads(branches_list)
        except: branches_list=[branches_list]

    with st.spinner("جاري تحميل البيانات..."):
        df_raw=fetch_outstanding_data()
    if df_raw.empty:
        st.info("📭 لا توجد بيانات متاحة"); return

    branch_col  = find_col(df_raw,['branch_name','Branch_name','اسم الفرع','branch','Branch'])
    officer_col = find_col(df_raw,['officer_name','Officer_name','اسم المسؤول','officer','Officer','emp_name','employee_name'])
    inst_col    = find_col(df_raw,['inst_amount','قيمة القسط','amount','inst_amt'])
    fawry_col   = find_col(df_raw,['fawry_amount','fawry_amt'])
    opay_col    = find_col(df_raw,['opay_amount','opay_amt','opayAmount'])
    client_col  = find_col(df_raw,['client_name','اسم العميل','client','name'])
    nation_col  = find_col(df_raw,['nation_id','الرقم القومي','nation','national_id'])

    if not branch_col:
        st.error("❌ لم يتم العثور على عمود الفرع"); return

    if not is_admin and branches_list:
        df_acc=df_raw[df_raw[branch_col].astype(str).isin(branches_list)].copy()
    else:
        df_acc=df_raw.copy()

    st.markdown('<div class="filter-bar"><div class="filter-bar-title">🔍 أدوات البحث والتصفية</div>',unsafe_allow_html=True)
    fc1,fc2,fc3,fc4=st.columns(4)

    with fc1:
        if not is_admin and branches_list:
            avail=sorted(df_acc[branch_col].dropna().unique().tolist())
            if len(avail)>1:
                sel_br=st.selectbox("🏢 الفرع",["الكل"]+avail,key="o_branch")
                if sel_br!="الكل": df_acc=df_acc[df_acc[branch_col]==sel_br]
            else:
                st.markdown(f"🏢 **الفرع:** {avail[0] if avail else '—'}")
        else:
            all_br=sorted(df_acc[branch_col].dropna().unique().tolist())
            sel_br=st.selectbox("🏢 الفرع",["الكل"]+all_br,key="o_branch_admin")
            if sel_br!="الكل": df_acc=df_acc[df_acc[branch_col]==sel_br]

    with fc2:
        off_f=None
        for c in df_acc.columns:
            if any(k in c.lower() for k in ['officer','مسؤول','مسئول']): off_f=c; break
        if off_f:
            offs=sorted(df_acc[off_f].dropna().unique().tolist())
            sel_off=st.selectbox("👤 الاخصائي",["الكل"]+offs,key="o_officer")
            if sel_off!="الكل": df_acc=df_acc[df_acc[off_f]==sel_off]

    with fc3:
        search_name=st.text_input("🔎 بحث باسم العميل",key="o_name")
        if search_name and client_col:
            df_acc=df_acc[df_acc[client_col].astype(str).str.contains(search_name,na=False,case=False)]

    with fc4:
        search_nat=st.text_input("🆔 بحث بالرقم القومي",key="o_nat")
        if search_nat and nation_col:
            df_acc=df_acc[df_acc[nation_col].astype(str).str.contains(search_nat,na=False,case=False)]

    st.markdown('</div>',unsafe_allow_html=True)

    if df_acc.empty:
        st.warning("⚠️ لا توجد بيانات تطابق معايير البحث"); return

    for c in [inst_col, fawry_col, opay_col]:
        if c: df_acc[c] = pd.to_numeric(df_acc[c], errors='coerce').fillna(0)

    fawry_series = df_acc[fawry_col] if fawry_col else pd.Series(0, index=df_acc.index)
    opay_series  = df_acc[opay_col]  if opay_col  else pd.Series(0, index=df_acc.index)
    inst_series  = df_acc[inst_col]  if inst_col  else pd.Series(0, index=df_acc.index)

    df_acc['_paid']      = fawry_series + opay_series
    df_acc['_inst']      = inst_series
    df_acc['_remaining'] = (df_acc['_inst'] - df_acc['_paid']).clip(lower=0)

    def get_status(row):
        paid = row['_paid']; inst = row['_inst']
        if inst <= 0:    return "❌ غير مدفوع"
        if paid >= inst: return "✅ مدفوع بالكامل"
        elif paid > 0:   return "⚠️ مسدد جزئي"
        else:            return "❌ غير مدفوع"

    df_acc['حالة الدفع'] = df_acc.apply(get_status, axis=1)
    status_order = {"✅ مدفوع بالكامل": 0, "⚠️ مسدد جزئي": 1, "❌ غير مدفوع": 2}
    df_acc['_sk'] = df_acc['حالة الدفع'].map(status_order)
    df_acc = df_acc.sort_values('_sk').drop('_sk', axis=1)

    ti = df_acc['_inst'].sum(); tp = df_acc['_paid'].sum(); tr = ti - tp
    df_partial  = df_acc[df_acc['حالة الدفع']=="⚠️ مسدد جزئي"]
    partial_cnt = len(df_partial); partial_paid=df_partial['_paid'].sum(); partial_rem=df_partial['_inst'].sum()-df_partial['_paid'].sum()
    df_unpaid   = df_acc[df_acc['حالة الدفع']=="❌ غير مدفوع"]; unpaid_cnt=len(df_unpaid)

    k1,k2,k3,k4 = st.columns(4)
    with k1:
        st.markdown(f'''<div class="kpi-card"><div class="kpi-lbl">💰 إجمالي المستحق</div><div class="kpi-val">{ti:,.0f} ج.م</div><div style="font-size:11px;color:#64748b;margin-top:4px">{len(df_acc):,} قسط</div></div>''', unsafe_allow_html=True)
    with k2:
        st.markdown(f'''<div class="kpi-card" style="border-top-color:#059669"><div class="kpi-lbl">💳 إجمالي المدفوع</div><div class="kpi-val" style="color:#059669">{tp:,.0f} ج.م</div><div style="font-size:11px;color:#059669;margin-top:4px">{len(df_acc[df_acc["حالة الدفع"]=="✅ مدفوع بالكامل"]):,} مكتمل</div></div>''', unsafe_allow_html=True)
    with k3:
        st.markdown(f'''<div class="kpi-card" style="border-top-color:#dc2626"><div class="kpi-lbl">📊 إجمالي المتبقي</div><div class="kpi-val" style="color:#dc2626">{tr:,.0f} ج.م</div><div style="font-size:11px;color:#dc2626;margin-top:4px">يشمل السداد الجزئي</div></div>''', unsafe_allow_html=True)
    with k4:
        partial_color="#f59e0b" if partial_cnt>0 else "#7c3aed"
        st.markdown(f'''<div class="kpi-card" style="border-top-color:{partial_color}"><div class="kpi-lbl">⚠️ مسدد جزئي</div><div class="kpi-val" style="color:{partial_color}">{partial_cnt:,} قسط</div><div style="font-size:11px;color:{partial_color};margin-top:4px">متبقي: {partial_rem:,.0f} ج.م</div></div>''', unsafe_allow_html=True)

    if partial_cnt > 0:
        st.markdown(f'''<div style="background:linear-gradient(90deg,#fffbeb,#fef3c7);border:1.5px solid #f59e0b;border-right:5px solid #f59e0b;border-radius:12px;padding:14px 20px;margin:14px 0;display:flex;align-items:center;gap:12px;"><div style="font-size:24px">⚠️</div><div><div style="font-weight:800;color:#92400e;font-size:14px">يوجد {partial_cnt:,} قسط مسدد جزئياً — لم يكتمل السداد بعد</div><div style="color:#b45309;font-size:12px;margin-top:3px">إجمالي ما تم سداده جزئياً: <strong>{partial_paid:,.2f} ج.م</strong> &nbsp;|&nbsp; إجمالي المتبقي منهم: <strong>{partial_rem:,.2f} ج.م</strong> &nbsp;|&nbsp; غير المدفوع كلياً: <strong>{unpaid_cnt:,} قسط</strong></div></div></div>''', unsafe_allow_html=True)

    unique_branches = df_acc[branch_col].dropna().unique().tolist()

    def build_branch_table(df_data, title_prefix=""):
        rows = []
        for br in sorted(df_data[branch_col].dropna().unique()):
            df_br = df_data[df_data[branch_col] == br]
            df_full=df_br[df_br['حالة الدفع']=="✅ مدفوع بالكامل"]; df_part=df_br[df_br['حالة الدفع']=="⚠️ مسدد جزئي"]; df_unp=df_br[df_br['حالة الدفع']=="❌ غير مدفوع"]
            inst_total=df_br['_inst'].sum(); paid_full=df_full['_paid'].sum(); paid_part=df_part['_paid'].sum(); remaining=inst_total-(paid_full+paid_part)
            rows.append({'اسم الفرع':br,'عدد الكل':len(df_br),'مسدد بالكامل (عدد)':len(df_full),'مسدد جزئياً (عدد)':len(df_part),'غير مدفوع (عدد)':len(df_unp),'إجمالي المستحق':inst_total,'إجمالي المسدد كلياً':paid_full,'إجمالي المسدد جزئياً':paid_part,'إجمالي المتبقي':remaining})
        if not rows: return
        df_tbl=pd.DataFrame(rows).sort_values('إجمالي المستحق',ascending=False)
        tot={'اسم الفرع':'الإجمالي الكلي','عدد الكل':df_tbl['عدد الكل'].sum(),'مسدد بالكامل (عدد)':df_tbl['مسدد بالكامل (عدد)'].sum(),'مسدد جزئياً (عدد)':df_tbl['مسدد جزئياً (عدد)'].sum(),'غير مدفوع (عدد)':df_tbl['غير مدفوع (عدد)'].sum(),'إجمالي المستحق':df_tbl['إجمالي المستحق'].sum(),'إجمالي المسدد كلياً':df_tbl['إجمالي المسدد كلياً'].sum(),'إجمالي المسدد جزئياً':df_tbl['إجمالي المسدد جزئياً'].sum(),'إجمالي المتبقي':df_tbl['إجمالي المتبقي'].sum()}
        def td(val,color='#1e293b',bold=False,bg=''):
            fw='font-weight:700;' if bold else 'font-weight:500;'; bgc=f'background:{bg};' if bg else ''
            return f"<td style='padding:10px 12px;text-align:center;border-bottom:1px solid #e2e8f0;color:{color};{fw}{bgc}'>{val}</td>"
        def td_r(val,color='#1e293b',bold=False):
            fw='font-weight:700;' if bold else 'font-weight:500;'
            return f"<td style='padding:10px 14px;text-align:right;border-bottom:1px solid #e2e8f0;color:{color};{fw}'>{val}</td>"
        def fmt(v): return f"{v:,.2f}"
        def fmti(v): return f"{int(v):,}"
        body=""
        for _,r in df_tbl.iterrows():
            body+=(f"<tr>"+td_r(r['اسم الفرع'],'#1e3a8a',True)+td(fmti(r['عدد الكل']),'#1e3a8a',True)+td(fmti(r['مسدد بالكامل (عدد)']),'#059669',True,'#f0fdf4')+td(fmti(r['مسدد جزئياً (عدد)']),'#d97706',True,'#fffbeb')+td(fmti(r['غير مدفوع (عدد)']),'#dc2626',True,'#fef2f2')+td(fmt(r['إجمالي المستحق']),'#1e3a8a')+td(fmt(r['إجمالي المسدد كلياً']),'#059669')+td(fmt(r['إجمالي المسدد جزئياً']),'#d97706')+td(fmt(r['إجمالي المتبقي']),'#dc2626',True)+f"</tr>")
        body+=(f"<tr style='background:#0f172a;'><td style='padding:12px 14px;text-align:right;color:#fff;font-weight:800;'>الإجمالي الكلي</td><td style='padding:12px;text-align:center;color:#fff;font-weight:800;'>{fmti(tot['عدد الكل'])}</td><td style='padding:12px;text-align:center;color:#6ee7b7;font-weight:800;'>{fmti(tot['مسدد بالكامل (عدد)'])}</td><td style='padding:12px;text-align:center;color:#fde68a;font-weight:800;'>{fmti(tot['مسدد جزئياً (عدد)'])}</td><td style='padding:12px;text-align:center;color:#fca5a5;font-weight:800;'>{fmti(tot['غير مدفوع (عدد)'])}</td><td style='padding:12px;text-align:center;color:#93c5fd;font-weight:800;'>{fmt(tot['إجمالي المستحق'])}</td><td style='padding:12px;text-align:center;color:#6ee7b7;font-weight:800;'>{fmt(tot['إجمالي المسدد كلياً'])}</td><td style='padding:12px;text-align:center;color:#fde68a;font-weight:800;'>{fmt(tot['إجمالي المسدد جزئياً'])}</td><td style='padding:12px;text-align:center;color:#fca5a5;font-weight:800;'>{fmt(tot['إجمالي المتبقي'])}</td></tr>")
        th_style="padding:13px 12px;text-align:center;color:#fff;font-size:13px;font-weight:700;white-space:nowrap;"
        th_style_r="padding:13px 14px;text-align:right;color:#fff;font-size:13px;font-weight:700;"
        header=(f"<tr style='background:#1e3a8a;'><th style='{th_style_r}'>اسم الفرع</th><th style='{th_style}'>عدد الكل</th><th style='{th_style}; background:#166534;'>✅ مسدد بالكامل</th><th style='{th_style}; background:#92400e;'>⚠️ مسدد جزئياً</th><th style='{th_style}; background:#991b1b;'>❌ غير مدفوع</th><th style='{th_style}'>إجمالي المستحق</th><th style='{th_style}; background:#166534;'>إجمالي المسدد كلياً</th><th style='{th_style}; background:#92400e;'>إجمالي المسدد جزئياً</th><th style='{th_style}; background:#991b1b;'>إجمالي المتبقي</th></tr>")
        st.markdown(f"<div style='overflow-x:auto;margin-bottom:20px;'><table style='width:100%;border-collapse:collapse;background:#fff;border-radius:14px;overflow:hidden;box-shadow:0 4px 20px rgba(30,58,138,0.12);font-size:13px;'><thead>{header}</thead><tbody>{body}</tbody></table></div>", unsafe_allow_html=True)

    if len(unique_branches)==1:
        br_name=unique_branches[0]
        st.markdown(f"<div style='background:#eff6ff;border-right:5px solid #2563eb;border-radius:10px;padding:10px 16px;margin-bottom:10px;font-size:15px;font-weight:700;color:#1e3a8a;'>📍 ملخص فرع: {br_name}</div>",unsafe_allow_html=True)
        build_branch_table(df_acc)
    else:
        st.markdown('<div class="sec-title">🏢 ملخص إجمالي الفروع</div>', unsafe_allow_html=True)
        build_branch_table(df_acc)

    st.markdown('<div class="sec-title">📋 جدول الأقساط المستحقة</div>',unsafe_allow_html=True)
    col_map=[('حالة الدفع','📊 حالة الدفع'),(branch_col,'🏢 اسم الفرع'),(client_col,'👤 اسم العميل'),(nation_col,'🆔 الرقم القومي'),('inst_mat_date','📅 تاريخ الاستحقاق'),(inst_col,'💰 قيمة القسط'),(fawry_col,'💳 مبلغ فوري'),(opay_col,'📱 مبلغ Opay'),(officer_col,'👨‍💼 المسؤول'),('loan_number','🔢 رقم القرض'),('inst_status','📌 حالة القسط')]
    dcols=[]; dnames=[]
    for col,name in col_map:
        if col and col in df_acc.columns: dcols.append(col); dnames.append(name)
    if dcols:
        df_disp=df_acc[dcols].copy(); df_disp.columns=dnames
        if '📅 تاريخ الاستحقاق' in df_disp.columns:
            df_disp['📅 تاريخ الاستحقاق']=pd.to_datetime(df_disp['📅 تاريخ الاستحقاق'],errors='coerce').dt.strftime('%Y-%m-%d')
        for c in ['💰 قيمة القسط','💳 مبلغ فوري','📱 مبلغ Opay']:
            if c in df_disp.columns:
                df_disp[c]=df_disp[c].apply(lambda x:f"{x:,.2f}" if pd.notna(x) and x>0 else "0.00")
        def color_rows(row):
            s=row['📊 حالة الدفع']
            if 'مسدد جزئي' in s: return ['background-color:#FFCCCC;color:#9C0006;font-weight:bold']*len(row)
            elif 'غير مدفوع' in s: return ['background-color:#FFE699;color:#7F4A00;font-weight:bold']*len(row)
            else: return ['background-color:#E2EFDA;color:#375623']*len(row)
        st.dataframe(df_disp.style.apply(color_rows,axis=1),use_container_width=True,height=450)

    if inst_col:
        with st.expander("📊 ملخص حسب حالة الدفع",expanded=False):
            sm=df_acc.groupby('حالة الدفع').agg({inst_col:'sum'}).reset_index()
            sm['عدد الأقساط']=df_acc.groupby('حالة الدفع').size().values
            sm['إجمالي فوري']=df_acc.groupby('حالة الدفع')[fawry_col].sum().values if fawry_col else 0
            sm['إجمالي Opay']=df_acc.groupby('حالة الدفع')[opay_col].sum().values if opay_col else 0
            sm['إجمالي المدفوع']=sm['إجمالي فوري']+sm['إجمالي Opay']
            paid_by_status=df_acc.groupby('حالة الدفع')['_paid'].sum().reset_index(); paid_by_status.columns=['حالة الدفع','_total_paid']
            sm=sm.merge(paid_by_status,on='حالة الدفع',how='left'); sm['المتبقي']=sm[inst_col]-sm['_total_paid'].fillna(0)
            sm=sm.drop(columns=['_total_paid']); sm.columns=['حالة الدفع','إجمالي المستحق','عدد الأقساط','إجمالي فوري','إجمالي Opay','إجمالي المدفوع','المتبقي']
            for c in ['إجمالي المستحق','إجمالي فوري','إجمالي Opay','إجمالي المدفوع','المتبقي']:
                sm[c]=sm[c].apply(lambda x:f"{x:,.2f}")
            sm['_o']=sm['حالة الدفع'].map(status_order); sm=sm.sort_values('_o').drop('_o',axis=1)
            st.dataframe(sm,use_container_width=True,hide_index=True)

    ofc=None
    for c in df_acc.columns:
        if any(k in c.lower() for k in ['officer','مسؤول','مسئول']): ofc=c; break
    if ofc and len(df_acc[ofc].dropna().unique())>0:
        with st.expander("👥 ملخص حسب المسؤول",expanded=False):
            try:
                os_=df_acc.groupby(ofc).agg({inst_col:['sum','count']}).reset_index(); os_.columns=['اسم المسؤول','إجمالي المستحق','عدد الأقساط']
                if fawry_col:
                    fs=df_acc.groupby(ofc)[fawry_col].sum().reset_index(); fs.columns=['اسم المسؤول','إجمالي فوري']
                    os_=os_.merge(fs,on='اسم المسؤول',how='left'); os_['إجمالي فوري']=os_['إجمالي فوري'].fillna(0)
                else: os_['إجمالي فوري']=0
                if opay_col:
                    ops=df_acc.groupby(ofc)[opay_col].sum().reset_index(); ops.columns=['اسم المسؤول','إجمالي Opay']
                    os_=os_.merge(ops,on='اسم المسؤول',how='left'); os_['إجمالي Opay']=os_['إجمالي Opay'].fillna(0)
                else: os_['إجمالي Opay']=0
                os_['إجمالي المدفوع']=os_['إجمالي فوري']+os_['إجمالي Opay']; os_['المتبقي']=os_['إجمالي المستحق']-os_['إجمالي المدفوع']; os_['متوسط القسط']=os_['إجمالي المستحق']/os_['عدد الأقساط']
                for c in ['إجمالي المستحق','إجمالي فوري','إجمالي Opay','إجمالي المدفوع','المتبقي','متوسط القسط']:
                    os_[c]=os_[c].apply(lambda x:f"{x:,.2f}")
                os_=os_.sort_values('إجمالي المستحق',ascending=False)
                st.dataframe(os_,use_container_width=True,hide_index=True)
            except Exception as e:
                st.warning(f"تعذر عرض ملخص المسؤولين: {e}")

    st.markdown('<div class="sec-title">📥 تحميل التقرير</div>',unsafe_allow_html=True)
    xls=generate_outstanding_excel(df_acc)
    st.download_button("📊 تحميل Excel ملون",data=xls,
        file_name=f"الاقساط_المستحقة_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ===================================================================
# ===================== الصفحة الرئيسية =====================
# ===================================================================

def main_app():
    if 'user' not in st.session_state:
        st.markdown("""
        <div style="background:linear-gradient(135deg,#0f172a 0%,#1e3a8a 100%);
                    border-radius:20px;padding:50px 20px;text-align:center;margin-bottom:30px;">
            <div style="font-size:56px;margin-bottom:10px">📊</div>
            <div style="font-size:28px;font-weight:800;color:white;margin-bottom:6px">نظام كاريتاس</div>
            <div style="color:#93c5fd;font-size:14px">لوحة التقارير والمتابعة</div>
        </div>
        """, unsafe_allow_html=True)
        c1,c2,c3=st.columns([1,1.4,1])
        with c2:
            with st.form("login_form"):
                st.markdown("### 🔐 تسجيل الدخول")
                username=st.text_input("👤 اسم المستخدم",placeholder="أدخل اسم المستخدم")
                password=st.text_input("🔑 كلمة المرور",type="password",placeholder="أدخل كلمة المرور")
                if st.form_submit_button("دخول",use_container_width=True):
                    user=check_login(username,password)
                    if user:
                        st.session_state['user']=user; st.rerun()
                    else:
                        st.error("❌ خطأ في اسم المستخدم أو كلمة المرور")
        return

    user=st.session_state['user']
    show_header(user)

    col_out1, col_out2 = st.columns([8, 2])
    with col_out2:
        if st.button("🚪 خروج", use_container_width=True):
            del st.session_state['user']
            st.rerun()

    user_id = user.get('id')
    unread_count = count_unread_notifications(user_id)
    
    if unread_count > 0:
        st.warning(f"🔔 لديك {unread_count} إشعار غير مقروء في نظام شكاوى العملاء")

    st.markdown(f"""
    <div style="text-align:center;margin:8px 0 28px;">
        <span style="font-size:15px;color:#64748b">
            مرحباً <strong style="color:#1e3a8a">{user['full_name']}</strong> — اختر الخدمة
        </span>
    </div>
    """, unsafe_allow_html=True)

    # ============================================================
    # استخدام st.expander و st.button فقط - بدون HTML معقد
    # ============================================================
    
    st.markdown("---")
    
    # الكارت الأول - سداد فوري
    with st.container():
        col1, col2 = st.columns([1, 4])
        with col1:
            st.markdown("# 💳")
        with col2:
            st.markdown("### سداد فوري & Opay")
            st.caption("عرض وتحليل بيانات السدادات — تقارير دقيقة ومتنوعة")
        if st.button("🔓 فتح سداد فوري", key="btn_reports", use_container_width=True):
            st.query_params["page"] = "reports"
            st.rerun()
    st.markdown("---")
    
    # الكارت الثاني - الأقساط المستحقة
    with st.container():
        col1, col2 = st.columns([1, 4])
        with col1:
            st.markdown("# 📋")
        with col2:
            st.markdown("### الأقساط المستحقة")
            st.caption("متابعة الأقساط المستحقة مع بيانات الدفع من فوري و Opay")
        if st.button("🔓 فتح الأقساط المستحقة", key="btn_installments", use_container_width=True):
            st.query_params["page"] = "installments"
            st.rerun()
    st.markdown("---")
    
    # الكارت الثالث - شكاوى العملاء (مع إشعار)
    with st.container():
        col1, col2 = st.columns([1, 4])
        with col1:
            if unread_count > 0:
                st.markdown("# 📝 🔴")
            else:
                st.markdown("# 📝")
        with col2:
            if unread_count > 0:
                st.markdown("### شكاوى العملاء ⚠️")
                st.caption(f"تسجيل ومتابعة شكاوى العملاء — إدخال وتعديل وموافقة إدارية **(لديك {unread_count} إشعار جديد)**")
            else:
                st.markdown("### شكاوى العملاء")
                st.caption("تسجيل ومتابعة شكاوى العملاء — إدخال وتعديل وموافقة إدارية")
        if st.button("🔓 فتح شكاوى العملاء", key="btn_complaints", use_container_width=True):
            st.query_params["page"] = "complaints"
            st.rerun()
    st.markdown("---")

    st.markdown("""
    <div style="text-align:center;margin-top:40px;padding:20px;
                color:#94a3b8;font-size:12px;border-top:1px solid #e2e8f0;">
        نظام كاريتاس للتقارير © 2025
    </div>
    """, unsafe_allow_html=True)


# ===================================================================
# ===================== تشغيل التطبيق =====================
# ===================================================================

query_params = st.query_params
page = query_params.get("page", "home")

if page != "home" and 'user' not in st.session_state:
    st.warning("⚠️ يجب تسجيل الدخول أولاً")
    if st.button("🔐 تسجيل الدخول"):
        st.query_params.clear(); st.rerun()
    st.stop()

if   page == "home":         main_app()
elif page == "reports":      reports_page()
elif page == "installments": outstanding_page()
elif page == "complaints":   complaints_page()
else:
    if 'user' in st.session_state: main_app()
    else: st.query_params.clear(); st.rerun()
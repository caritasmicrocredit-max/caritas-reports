import streamlit as st
import pandas as pd
from supabase import create_client
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="سداد فورى و opay",
    page_icon="💳",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;600;700;900&display=swap');
* { font-family: 'Cairo', sans-serif !important; }
.main { direction: rtl; text-align: right; }
[data-testid="stSidebar"] { direction: rtl; background-color: #f8f9fa; }
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
.stDeployButton {display:none;}
input { text-align: right; direction: rtl; }

.page-header {
    background: linear-gradient(135deg, #1e3a8a 0%, #1d4ed8 100%);
    border-radius: 18px;
    padding: 24px 36px;
    margin-bottom: 28px;
    color: white;
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.page-header-title { font-size: 24px; font-weight: 900; margin: 0 0 4px 0; }
.page-header-sub { font-size: 13px; opacity: 0.8; margin: 0; }
.page-header-icon { font-size: 52px; }

.metric-card {
    background: white;
    padding: 20px;
    border-radius: 12px;
    box-shadow: 0 4px 16px rgba(0,0,0,0.07);
    border-top: 5px solid #1e3a8a;
    text-align: center;
}
.metric-value { font-size: 22px; font-weight: bold; color: #1e3a8a; }
.metric-label { font-size: 14px; color: #6b7280; margin-bottom: 6px; }
.stDataFrame { border: 1px solid #e5e7eb; border-radius: 10px; }

/* ===== Login ===== */
.login-wrap {
    display: flex; justify-content: center;
    padding: 60px 0;
}
.login-box {
    background: white;
    padding: 40px 36px;
    border-radius: 20px;
    box-shadow: 0 12px 40px rgba(30,58,138,0.15);
    border: 1.5px solid #bfdbfe;
    text-align: center;
    width: 100%;
}
.login-icon { font-size: 56px; }
.login-title { color:#1e3a8a; font-size:20px; font-weight:900; margin:10px 0 4px 0; }
.login-sub { color:#6b7280; font-size:13px; margin:0 0 24px 0; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# Supabase
# ============================================================
URL = st.secrets["SUPABASE_URL"]
KEY = st.secrets["SUPABASE_KEY"]
supabase = create_client(URL, KEY)

SERVICE_CODES = ["75146", "90074"]

# ============================================================
# وظائف مساعدة
# ============================================================
def check_login(u, p):
    res = supabase.table("app_users").select("*").eq("id", u).eq("password_hash", p).execute()
    return res.data[0] if res.data else None

def thin_border():
    s = Side(border_style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def write_total_row(ws, total_row, cols, last_data_row):
    DARK_BLUE = "1E3A8A"
    TOTAL_BG  = "BFDBFE"
    for ci in range(1, len(cols) + 1):
        c = ws.cell(row=total_row, column=ci)
        c.fill = PatternFill("solid", fgColor=TOTAL_BG)
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(bold=True, color=DARK_BLUE, name="Arial", size=11)
    ws.cell(row=total_row, column=1).value = "✦ الإجمالي"
    if 'المبلغ' in cols:
        amt_ci = cols.index('المبلغ') + 1
        col_letter = get_column_letter(amt_ci)
        ws.cell(row=total_row, column=amt_ci).value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
        ws.cell(row=total_row, column=amt_ci).number_format = '#,##0.00'
    ws.row_dimensions[total_row].height = 26

def fetch_data(service_codes, is_admin, user_branches):
    all_data = []
    limit, offset = 1000, 0
    while True:
        res = (supabase.table("all_payments_report")
               .select("*")
               .in_("كود الخدمة", service_codes)
               .range(offset, offset + limit - 1)
               .execute())
        if not res.data:
            break
        all_data.extend(res.data)
        if len(res.data) < limit:
            break
        offset += limit
    df = pd.DataFrame(all_data)
    if not df.empty:
        df['تاريخ الدفع'] = pd.to_datetime(df['تاريخ الدفع'], dayfirst=True, errors='coerce')
        if not is_admin:
            df = df[df['branch_name'].isin(user_branches)]
    return df

def generate_excel_single(df_display, sheet_title="التقرير", report_title="تقرير السدادات"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.sheet_view.rightToLeft = True
    DARK_BLUE  = "1E3A8A"
    LIGHT_BLUE = "EFF6FF"
    ALT_ROW    = "F0F4FF"
    WHITE      = "FFFFFF"
    cols = list(df_display.columns)
    last_col = get_column_letter(len(cols))
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].value = report_title
    ws['A1'].font = Font(bold=True, size=14, color=DARK_BLUE, name="Arial")
    ws['A1'].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32
    for ci, header in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=header)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=11)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[2].height = 24
    for ri, row in enumerate(df_display.itertuples(index=False), 3):
        bg = ALT_ROW if ri % 2 == 0 else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.font = Font(name="Arial", size=10)
            c.border = thin_border()
    last_data_row = 2 + len(df_display)
    write_total_row(ws, last_data_row + 1, cols, last_data_row)
    col_widths = {"اسم العميل": 28, "الفرع": 20}
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 18)
    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def generate_excel_daily(df_display, original_df):
    wb = Workbook()
    wb.remove(wb.active)
    DARK_BLUE  = "1E3A8A"
    LIGHT_BLUE = "EFF6FF"
    ALT_ROW    = "F0F4FF"
    TOTAL_BG   = "BFDBFE"
    WHITE      = "FFFFFF"

    def style_sheet(ws, df_part, title_text):
        ws.sheet_view.rightToLeft = True
        cols = list(df_part.columns)
        last_col = get_column_letter(len(cols))
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'].value = title_text
        ws['A1'].font = Font(bold=True, size=13, color=DARK_BLUE, name="Arial")
        ws['A1'].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=DARK_BLUE)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border()
        ws.row_dimensions[2].height = 22
        for ri, row in enumerate(df_part.itertuples(index=False), 3):
            bg = ALT_ROW if ri % 2 == 0 else WHITE
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.font = Font(name="Arial", size=10)
                c.border = thin_border()
        last_data_row = 2 + len(df_part)
        write_total_row(ws, last_data_row + 1, cols, last_data_row)
        ws.freeze_panes = "A3"
        col_widths = {"اسم العميل": 28, "الفرع": 20}
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 17)

    ws_sum = wb.create_sheet("ملخص يومي")
    ws_sum.sheet_view.rightToLeft = True
    summary_cols = ["التاريخ", "عدد العمليات", "إجمالي المبلغ (ج.م)"]
    n_sc = len(summary_cols)
    ws_sum.merge_cells(f'A1:{get_column_letter(n_sc)}1')
    ws_sum['A1'].value = "ملخص يومي - تقرير السدادات"
    ws_sum['A1'].font = Font(bold=True, size=14, color=DARK_BLUE, name="Arial")
    ws_sum['A1'].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws_sum['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 32
    for ci, h in enumerate(summary_cols, 1):
        c = ws_sum.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color=WHITE, name="Arial")
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()
    ws_sum.row_dimensions[2].height = 22
    temp = original_df.copy()
    temp['_date'] = temp['تاريخ الدفع'].dt.date
    dates_sorted = sorted(temp['_date'].dropna().unique())
    for ri, d in enumerate(dates_sorted, 3):
        day_df = temp[temp['_date'] == d]
        bg = ALT_ROW if ri % 2 == 0 else WHITE
        for ci in range(1, n_sc + 1):
            c = ws_sum.cell(row=ri, column=ci)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.font = Font(name="Arial", size=10)
            c.border = thin_border()
        ws_sum.cell(row=ri, column=1).value = str(d)
        ws_sum.cell(row=ri, column=2).value = len(day_df)
        ws_sum.cell(row=ri, column=3).value = float(day_df['المبلغ'].sum())
        ws_sum.cell(row=ri, column=3).number_format = '#,##0.00'
    total_row_sum = 2 + len(dates_sorted) + 1
    for ci in range(1, n_sc + 1):
        c = ws_sum.cell(row=total_row_sum, column=ci)
        c.fill = PatternFill("solid", fgColor=TOTAL_BG)
        c.border = thin_border()
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.font = Font(bold=True, color=DARK_BLUE, name="Arial")
    ws_sum.cell(row=total_row_sum, column=1).value = "✦ الإجمالي الكلي"
    ws_sum.cell(row=total_row_sum, column=3).value = f"=SUM(C3:C{total_row_sum - 1})"
    ws_sum.cell(row=total_row_sum, column=3).number_format = '#,##0.00'
    ws_sum.row_dimensions[total_row_sum].height = 26
    for ci, w in enumerate([18, 18, 25], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.freeze_panes = "A3"
    for d in dates_sorted:
        day_df_orig = temp[temp['_date'] == d].copy()
        day_display = day_df_orig.rename(columns={
            'client_code': 'كود العميل',
            'client_name': 'اسم العميل',
            'branch_name': 'الفرع'
        })
        if 'تاريخ الدفع' in day_display.columns:
            day_display['تاريخ الدفع'] = day_display['تاريخ الدفع'].dt.strftime('%Y-%m-%d')
        drop_cols = [c for c in day_display.columns if c.startswith('_') or c == 'id']
        day_display = day_display.drop(columns=drop_cols, errors='ignore')
        ws_day = wb.create_sheet(str(d)[:31])
        style_sheet(ws_day, day_display, f"تقرير سدادات يوم {d}")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
#                      الواجهة
# ============================================================

# ===== تسجيل الدخول =====
if 'user_sadad' not in st.session_state:
    st.markdown("""
    <div style="text-align:center; padding:50px 0 20px;">
        <div style="font-size:60px;">💳</div>
        <h1 style="color:#1e3a8a;font-size:24px;font-weight:900;margin:12px 0 4px;">
            متابعة سداد فورى و opay
        </h1>
        <p style="color:#6b7280;font-size:13px;margin:0 0 30px;">
            أكواد الخدمة: 75146 + 90074
        </p>
    </div>
    """, unsafe_allow_html=True)

    cc1, cc2, cc3 = st.columns([1, 1.2, 1])
    with cc2:
        with st.form("login_sadad"):
            st.markdown("""
            <div style="background:white;padding:28px;border-radius:18px;
                 box-shadow:0 8px 32px rgba(30,58,138,0.12);border:1.5px solid #bfdbfe;">
            """, unsafe_allow_html=True)
            u = st.text_input("👤 اسم المستخدم")
            p = st.text_input("🔒 كلمة المرور", type="password")
            login_btn = st.form_submit_button("🚀 دخول", use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)

            if login_btn:
                if not u or not p:
                    st.error("يرجى إدخال البيانات كاملة")
                else:
                    user = check_login(u, p)
                    if user:
                        st.session_state['user_sadad'] = user
                        st.rerun()
                    else:
                        st.error("❌ بيانات الدخول غير صحيحة")

# ===== لوحة التحكم =====
else:
    user = st.session_state['user_sadad']
    is_admin = user.get('role') == 'admin'
    user_branches = user.get('branches', [])

    # السايدبار
    st.sidebar.markdown(f"""
    <div style="text-align:center;padding:12px 0 6px;">
        <div style="font-size:36px;">💳</div>
        <h3 style="color:#1e3a8a;margin:6px 0 2px;font-size:15px;">سداد فورى و opay</h3>
        <p style="color:#6b7280;font-size:12px;margin:0;">{user['full_name']}</p>
    </div>
    """, unsafe_allow_html=True)
    st.sidebar.divider()

    if st.sidebar.button("🚪 خروج", use_container_width=True):
        del st.session_state['user_sadad']
        st.rerun()

    st.sidebar.divider()
    st.sidebar.header("🔍 أدوات البحث")
    s_name = st.sidebar.text_input("بحث باسم العميل")
    s_code = st.sidebar.text_input("بحث بكود العميل")

    # جلب البيانات
    df_raw = fetch_data(SERVICE_CODES, is_admin, user_branches)

    if df_raw.empty:
        st.warning("⚠️ لا توجد بيانات لأكواد الخدمة المحددة.")
        st.stop()

    v_dates = df_raw['تاريخ الدفع'].dropna()
    start_d = st.sidebar.date_input("من تاريخ", v_dates.min().date())
    end_d   = st.sidebar.date_input("إلى تاريخ", v_dates.max().date())

    codes_available = ["الكل"] + sorted(df_raw['كود الخدمة'].unique().tolist())
    sel_code = st.sidebar.selectbox("كود الخدمة", codes_available)

    # الفلترة
    mask = (df_raw['تاريخ الدفع'].dt.date >= start_d) & (df_raw['تاريخ الدفع'].dt.date <= end_d)
    if sel_code != "الكل":
        mask &= (df_raw['كود الخدمة'] == sel_code)
    if s_name:
        mask &= df_raw['client_name'].astype(str).str.contains(s_name, na=False, case=False)
    if s_code:
        mask &= df_raw['client_code'].astype(str).str.contains(s_code, na=False, case=False)

    final_df = df_raw.loc[mask]

    # الهيدر
    st.markdown(f"""
    <div class="page-header">
        <div>
            <p class="page-header-title">💳 متابعة سداد فورى و opay</p>
            <p class="page-header-sub">
                أكواد: {' + '.join(SERVICE_CODES)} &nbsp;|&nbsp;
                الفترة: {start_d} → {end_d} &nbsp;|&nbsp;
                المستخدم: {user['full_name']}
            </p>
        </div>
        <div class="page-header-icon">📑</div>
    </div>
    """, unsafe_allow_html=True)

    # الكروت
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">إجمالي المبالغ</div>
            <div class="metric-value">{final_df['المبلغ'].sum():,.2f} ج.م</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">عدد العمليات</div>
            <div class="metric-value">{len(final_df)} حركة</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        counts = final_df['كود الخدمة'].value_counts()
        codes_html = "".join([f"<div>{k}: {v}</div>" for k, v in counts.items()])
        st.markdown(f"""<div class="metric-card">
            <div class="metric-label">تفاصيل الأكواد</div>
            <div style="font-weight:bold;color:#1e3a8a;">{codes_html}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # الملخص اليومي
    with st.expander("📅 عرض الملخص اليومي", expanded=False):
        daily_summary = (
            final_df.groupby(final_df['تاريخ الدفع'].dt.date)
            .agg(عدد_العمليات=('المبلغ', 'count'), إجمالي_المبلغ=('المبلغ', 'sum'))
            .reset_index()
        )
        daily_summary.columns = ['التاريخ', 'عدد العمليات', 'إجمالي المبلغ (ج.م)']
        st.dataframe(daily_summary, use_container_width=True, hide_index=True)

    # الجدول الرئيسي
    display_df = final_df.copy().rename(columns={
        'client_code': 'كود العميل',
        'client_name': 'اسم العميل',
        'branch_name': 'الفرع'
    })
    drop_cols = [c for c in display_df.columns if c.startswith('_') or c == 'id']
    display_df = display_df.drop(columns=drop_cols, errors='ignore')
    if 'تاريخ الدفع' in display_df.columns:
        display_df['تاريخ الدفع'] = display_df['تاريخ الدفع'].dt.strftime('%Y-%m-%d')

    st.dataframe(display_df, use_container_width=True, hide_index=True)

    # التحميل
    st.sidebar.divider()
    st.sidebar.header("📥 تحميل التقرير")

    split_mode = st.sidebar.radio(
        "نوع التنزيل",
        ["كل البيانات في شيت واحد", "تقسيم يوم يوم (شيت لكل يوم)"],
        index=0
    )

    if split_mode == "تقسيم يوم يوم (شيت لكل يوم)":
        available_dates = sorted(final_df['تاريخ الدفع'].dropna().dt.date.unique())
        date_options = ["كل الأيام"] + [str(d) for d in available_dates]
        selected_day = st.sidebar.selectbox("اختر اليوم للتنزيل", date_options)

        if selected_day == "كل الأيام":
            excel_bytes = generate_excel_daily(display_df, final_df)
            file_label = f"سداد_فوري_كل_الأيام_{datetime.now().date()}.xlsx"
        else:
            sel_date = pd.to_datetime(selected_day).date()
            day_display = display_df[
                pd.to_datetime(display_df['تاريخ الدفع'], errors='coerce').dt.date == sel_date
            ]
            excel_bytes = generate_excel_single(
                day_display,
                sheet_title=selected_day,
                report_title=f"تقرير سدادات يوم {selected_day}"
            )
            file_label = f"سداد_فوري_{selected_day}.xlsx"
    else:
        excel_bytes = generate_excel_single(
            display_df,
            report_title=f"تقرير سداد فورى و opay — {start_d} إلى {end_d}"
        )
        file_label = f"سداد_فوري_{datetime.now().date()}.xlsx"

    st.sidebar.download_button(
        label="📊 تحميل Excel ملون",
        data=excel_bytes,
        file_name=file_label,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
